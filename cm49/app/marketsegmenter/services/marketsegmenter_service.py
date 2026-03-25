from __future__ import annotations

import csv
import difflib
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook

from core.config_loader import list_country_config_files, load_csv_rows, load_yaml_config

ProgressCallback = Callable[[int, str], None]
LogCallback = Callable[[str], None]

_MS_MAPPING = load_yaml_config('marketsegmenter/mapping_fields.yaml')
_MS_ALIASES = load_yaml_config('marketsegmenter/column_aliases.yaml')
_MS_LANGS = load_yaml_config('marketsegmenter/country_langs.yaml')
_MS_NAMES = load_yaml_config('marketsegmenter/country_name_to_code.yaml')
_MS_FAMILY = load_yaml_config('marketsegmenter/family_scoring.yaml')
_MS_PRICE = load_yaml_config('marketsegmenter/price_rules.yaml')
_MS_GLOBAL_KEYWORDS = load_yaml_config('marketsegmenter/global_keywords.yaml')
_MS_NEGATIVE = load_yaml_config('marketsegmenter/negative_keywords.yaml')

MARKETSEGMENTER_MAPPING_FIELDS = list(_MS_MAPPING.get('mapping_fields', []))
MARKETSEGMENTER_REQUIRED_FIELDS = set(_MS_MAPPING.get('required_fields', []))
MARKETSEGMENTER_COLUMN_ALIASES = _MS_ALIASES.get('column_aliases', {})
COUNTRY_LANGS = _MS_LANGS.get('country_langs', {})
COUNTRY_NAME_TO_CODE = _MS_NAMES.get('country_name_to_code', {})
ALL_TEXT_FIELDS = list(_MS_MAPPING.get('all_text_fields', []))
DEBUG_OUTPUT_COLUMNS = list(_MS_MAPPING.get('debug_output_columns', []))
BASE_FAMILY_SCORE_RULES = {
    (row.get('segment0', ''), row.get('segment1', '')): float(row.get('initial_score', 0.0))
    for row in _MS_FAMILY.get('base_family_score_rules', [])
}
PRICE_BUCKETS = _MS_PRICE.get('price_buckets', {})
PRICE_SCORE_RULES = {
    bucket: [((item.get('segment0', ''), item.get('segment1', '')), float(item.get('delta', 0.0)), item.get('reason', '')) for item in items]
    for bucket, items in _MS_PRICE.get('price_score_rules', {}).items()
}
KEYWORD_RULES = list(_MS_GLOBAL_KEYWORDS.get('keyword_rules', []))
NEGATIVE_KEYWORDS = list(_MS_NEGATIVE.get('negative_keywords', []))
COUNTRY_KEYWORD_RULES: list[dict] = []
for cfg_path in list_country_config_files('marketsegmenter'):
    payload = load_yaml_config(f'marketsegmenter/countries/{cfg_path.name}')
    code = str(payload.get('country_code', '')).upper()
    for rule in payload.get('keyword_rules', []):
        COUNTRY_KEYWORD_RULES.append({'countries': [code], **rule})


def _noop_progress(percent: int, message: str) -> None:
    return None


def _noop_log(message: str) -> None:
    return None


def _normalize_text(value: object) -> str:
    text = '' if value is None else str(value)
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = re.sub(r"[\u2019'`´]+", ' ', text)
    text = re.sub(r'[^a-z0-9&/+ ]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _split_csv_like(raw: str) -> list[str]:
    raw = (raw or '').strip()
    return [part.strip() for part in re.split(r'[,\|;/]+', raw) if part.strip()] if raw else []


def suggest_column_mapping(columns: list[str]) -> dict[str, str]:
    normalized_to_original = {_normalize_text(col).replace(' ', '_'): col for col in columns}
    suggestions: dict[str, str] = {}
    used: set[str] = set()
    for target, aliases in MARKETSEGMENTER_COLUMN_ALIASES.items():
        for alias in aliases:
            candidate = normalized_to_original.get(_normalize_text(alias).replace(' ', '_'))
            if candidate and candidate not in used:
                suggestions[target] = candidate
                used.add(candidate)
                break
        if target in suggestions:
            continue
        for original in columns:
            if original in used:
                continue
            n = _normalize_text(original)
            if any(_normalize_text(alias) in n or n in _normalize_text(alias) for alias in aliases):
                suggestions[target] = original
                used.add(original)
                break
    return suggestions



def _detect_csv_delimiter(path: Path) -> str:
    sample = path.read_text(encoding='utf-8-sig', errors='ignore')[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;	|')
        return dialect.delimiter
    except Exception:
        return ','


def _iter_source_rows(path: Path, sheet_name: str | None = None):
    suffix = path.suffix.lower()
    if suffix in {'.csv', '.txt'}:
        delimiter = _detect_csv_delimiter(path)
        with path.open('r', encoding='utf-8-sig', newline='') as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            headers = list(reader.fieldnames or [])
            for row in reader:
                yield headers, {str(k): '' if v is None else str(v) for k, v in row.items()}
        return
    if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            wb.close()
            return
        headers = ['' if v is None else str(v) for v in raw_headers]
        for values in rows:
            row = {}
            for idx, header in enumerate(headers):
                value = values[idx] if idx < len(values) else ''
                row[header] = '' if value is None else str(value)
            yield headers, row
        wb.close()
        return
    raise ValueError(f'Format non supporté pour le market segmenter: {path.suffix}')


def _estimate_total_rows(path: Path, sheet_name: str | None = None) -> int | None:
    suffix = path.suffix.lower()
    try:
        if suffix in {'.csv', '.txt'}:
            with path.open('r', encoding='utf-8-sig', errors='ignore') as fh:
                total = sum(1 for _ in fh)
            return max(total - 1, 0)
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            total = max((ws.max_row or 1) - 1, 0)
            wb.close()
            return total
    except Exception:
        return None
    return None


def _prepare_output_headers(source_headers: list[str], mapping: dict[str, str]) -> list[str]:
    reverse = {source: target for target, source in mapping.items() if source in source_headers}
    output_headers = [reverse.get(col, col) for col in source_headers]
    existing = set(output_headers)
    for field in MARKETSEGMENTER_MAPPING_FIELDS:
        if field not in existing:
            output_headers.append(field)
            existing.add(field)
    return output_headers


def _map_row_dict(row: dict[str, str], source_headers: list[str], mapping: dict[str, str], output_headers: list[str]) -> dict[str, str]:
    reverse = {source: target for target, source in mapping.items() if source in source_headers}
    out: dict[str, str] = {}
    for col in source_headers:
        out[reverse.get(col, col)] = '' if row.get(col) is None else str(row.get(col, ''))
    for field in output_headers:
        out.setdefault(field, '')
    return out


def _csv_safe(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = str(value)
    return text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
def inspect_marketsegmenter_file(uploaded_file) -> dict:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
        xls = pd.ExcelFile(uploaded_file)
        sheets = []
        for name in xls.sheet_names[:10]:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, sheet_name=name, nrows=20)
            columns = [str(c) for c in df.columns]
            suggestions = suggest_column_mapping(columns)
            sheets.append({
                'name': name,
                'max_row': None,
                'max_column': len(columns),
                'preview': [columns] + df.head(20).fillna('').astype(str).values.tolist(),
                'detected_columns': columns,
                'mapping_suggestions': suggestions,
                'missing_required': sorted(MARKETSEGMENTER_REQUIRED_FIELDS - set(suggestions.keys())),
            })
        return {'filename': Path(uploaded_file.name).name, 'kind': 'excel', 'sheets': sheets}
    if suffix in {'.csv', '.txt'}:
        df = pd.read_csv(uploaded_file, nrows=20)
        columns = [str(c) for c in df.columns]
        suggestions = suggest_column_mapping(columns)
        return {
            'filename': Path(uploaded_file.name).name,
            'kind': 'csv',
            'sheets': [{
                'name': '__csv__',
                'max_row': None,
                'max_column': len(columns),
                'preview': [columns] + df.head(20).fillna('').astype(str).values.tolist(),
                'detected_columns': columns,
                'mapping_suggestions': suggestions,
                'missing_required': sorted(MARKETSEGMENTER_REQUIRED_FIELDS - set(suggestions.keys())),
            }],
        }
    raise ValueError('Inspection market segmenter disponible uniquement pour CSV et Excel.')


@dataclass
class MarketSegmenterOptions:
    marketsegmenter_sheet_name: str | None = None
    marketsegmenter_mapping: dict[str, str] = field(default_factory=dict)
    country_default: str = ''
    enable_typo_tolerance: bool = True
    emit_debug_columns: bool = True


class MarketSegmenterService:
    def __init__(self, progress_callback: ProgressCallback | None = None, log_callback: LogCallback | None = None):
        self.progress_callback = progress_callback or _noop_progress
        self.log_callback = log_callback or _noop_log
        self.type_mapping_df = self._load_type_mapping_df()
        self.type_mapping = {
            _normalize_text(row['type']): row
            for _, row in self.type_mapping_df.fillna('').iterrows()
            if str(row.get('type', '')).strip()
        }
        self.compiled_keyword_rules = self._compile_keyword_rules(KEYWORD_RULES)
        self.compiled_country_keyword_rules = self._compile_keyword_rules(COUNTRY_KEYWORD_RULES)
        self.compiled_negative_keyword_rules = self._compile_keyword_rules(NEGATIVE_KEYWORDS)

    def progress(self, percent: int, message: str) -> None:
        self.progress_callback(percent, message)

    def log(self, message: str) -> None:
        self.log_callback(message)

    def run(self, input_path: Path, output_path: Path, options: MarketSegmenterOptions) -> Path:
        self.progress(5, 'Chargement du fichier source du market segmenter')
        output_path.parent.mkdir(parents=True, exist_ok=True)

        debug_cols = DEBUG_OUTPUT_COLUMNS if options.emit_debug_columns else []
        stats = defaultdict(int)
        total = _estimate_total_rows(input_path, options.marketsegmenter_sheet_name) or 0
        processed = 0
        source_headers: list[str] | None = None
        projected_headers: list[str] | None = None

        with output_path.open('w', newline='', encoding='utf-8-sig') as fh:
            writer = None
            for raw_headers, raw_row in _iter_source_rows(input_path, options.marketsegmenter_sheet_name):
                if source_headers is None:
                    source_headers = list(raw_headers)
                    projected_headers = _prepare_output_headers(source_headers, options.marketsegmenter_mapping)
                    output_columns = list(projected_headers) + [
                        'fyre_market_segment_type0', 'fyre_market_segment_type1',
                        'fyre_market_segment_type2', 'fyre_market_segment_type3',
                        'segmentation_confidence', 'segmentation_reasons',
                        'base_main_type_path', 'all_types_paths_considered',
                        'keyword_hits', 'language_scope',
                    ] + debug_cols
                    writer = csv.DictWriter(
                        fh,
                        fieldnames=output_columns,
                        delimiter=';',
                        quotechar='"',
                        quoting=csv.QUOTE_ALL,
                        lineterminator='\n',
                    )
                    writer.writeheader()

                row = _map_row_dict(raw_row, source_headers or [], options.marketsegmenter_mapping, projected_headers or [])
                classified = self._classify_row(row, options)
                out = {col: _csv_safe(row.get(col, '')) for col in (projected_headers or [])}
                out.update({k: _csv_safe(v) for k, v in classified.items()})
                writer.writerow(out)
                processed += 1
                stats[classified.get('fyre_market_segment_type0') or 'unknown'] += 1
                if processed == 1 or processed % 1000 == 0 or (total and processed == total):
                    self.progress(min(96, 15 + int((processed / max(total, processed)) * 80)), f'Segmentation en cours : {processed}/{total or "?"}')

        summary = {
            'rows': int(processed),
            'macro_stats': dict(stats),
            'type_mapping_coverage': int(self.type_mapping_df['marketsegment0'].astype(str).str.strip().ne('').sum()),
            'typo_tolerance': bool(options.enable_typo_tolerance),
            'country_default': options.country_default or '',
            'country_scope_enabled': True,
            'price_signal_enabled': True,
            'culinary_keywords_enabled': True,
            'streaming_mode_enabled': True,
            'safe_csv_export': True,
        }
        output_path.with_name(output_path.stem + '_summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
        self.progress(100, 'Market segmenter terminé')
        return output_path

    def _load_type_mapping_df(self) -> pd.DataFrame:
        mapping_path = Path(__file__).resolve().parents[2] / 'config_catalog' / 'marketsegmenter' / 'type_mapping.csv'
        if mapping_path.exists():
            return pd.read_csv(mapping_path).fillna('')
        return pd.DataFrame(columns=['type', 'marketsegment0', 'marketsegment1', 'marketsegment2', 'marketsegment3', 'mapping_reason'])

    def _apply_mapping(self, df: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
        reverse = {source: target for target, source in mapping.items() if source in df.columns}
        df = df.rename(columns=reverse)
        for field in MARKETSEGMENTER_MAPPING_FIELDS:
            if field not in df.columns:
                df[field] = ''
            df[field] = df[field].fillna('').astype(str)
        return df

    def _compile_keyword_rules(self, rules: list[dict]) -> list[dict]:
        compiled = []
        for rule in rules:
            normalized_keywords = []
            for kw in rule.get('keywords', []):
                nkw = _normalize_text(kw)
                if nkw:
                    normalized_keywords.append(nkw)
            compiled.append({**rule, 'normalized_keywords': tuple(dict.fromkeys(normalized_keywords))})
        return compiled

    def _classify_row(self, row: pd.Series, options: MarketSegmenterOptions) -> dict:
        country_code = self._resolve_country_code(row, options)
        langs = COUNTRY_LANGS.get(country_code, ['en'])
        haystack = _normalize_text(' | '.join(str(row.get(k, '')) for k in ALL_TEXT_FIELDS))
        fuzzy_haystack = _normalize_text(' | '.join(str(row.get(k, '')) for k in ['name', 'main_type', 'all_types', 'description_1', 'description_2', 'description_3']))
        tokens = haystack.split()
        fuzzy_tokens = tuple(sorted(set(fuzzy_haystack.split())))

        detailed_scores: dict[tuple[str, str, str, str], float] = defaultdict(float)
        family_scores: dict[tuple[str, str], float] = dict(BASE_FAMILY_SCORE_RULES)
        reasons: list[str] = []
        keyword_hits: list[str] = []
        cuisine_hits: list[str] = []

        main_type_path, main_reason = self._map_google_type(row.get('main_type', ''))
        if any(main_type_path):
            detailed_scores[main_type_path] += 100.0
            family_scores[(main_type_path[0], main_type_path[1])] = family_scores.get((main_type_path[0], main_type_path[1]), 0.0) + 16.0
            reasons.append(f"main_type:{row.get('main_type', '')}=>{' > '.join([p for p in main_type_path if p])}")
            if main_reason:
                reasons.append(f"main_type_reason:{main_reason}")

        all_types_paths: list[str] = []
        for idx, candidate_type in enumerate(_split_csv_like(row.get('all_types', ''))[:25], start=1):
            path, reason = self._map_google_type(candidate_type)
            if any(path):
                same_family = path[:2] == main_type_path[:2]
                detailed_scores[path] += 45.0 if same_family else 70.0
                family_scores[(path[0], path[1])] = family_scores.get((path[0], path[1]), 0.0) + (8.0 if same_family else 12.0)
                all_types_paths.append(' > '.join([p for p in path if p]))
                reasons.append(f"all_types[{idx}]:{candidate_type}=>{' > '.join([p for p in path if p])}")
                if reason:
                    reasons.append(f"all_types_reason[{idx}]:{reason}")

        ordered_pre = sorted(family_scores.values(), reverse=True)
        prelim_gap = ordered_pre[0] - ordered_pre[1] if len(ordered_pre) > 1 else (ordered_pre[0] if ordered_pre else 0.0)
        deep_scan = prelim_gap < 18.0

        for rule in self.compiled_keyword_rules:
            hits = self._match_rule_hits(haystack, fuzzy_tokens, rule, options.enable_typo_tolerance and deep_scan)
            if hits:
                segment = tuple((rule.get('segment') or ['', '', '', ''])[:4])
                detailed_scores[segment] += float(rule.get('weight', 10))
                family = tuple(rule.get('family') or segment[:2])
                family_scores[family] = family_scores.get(family, 0.0) + float(rule.get('weight', 10)) * 0.65
                keyword_hits.extend(hits[:3])
                cuisine_hits.extend(hits[:3])
                reasons.append(f"keywords:{'/'.join(hits[:3])}=>{' > '.join([p for p in segment if p])}")

        for rule in self.compiled_country_keyword_rules:
            if country_code and country_code not in set(rule.get('countries', [])):
                continue
            hits = self._match_rule_hits(haystack, fuzzy_tokens, rule, options.enable_typo_tolerance and deep_scan)
            if hits:
                segment = tuple((rule.get('segment') or ['', '', '', ''])[:4])
                detailed_scores[segment] += float(rule.get('weight', 10))
                family = tuple(rule.get('family') or segment[:2])
                family_scores[family] = family_scores.get(family, 0.0) + float(rule.get('weight', 10)) * 0.7
                keyword_hits.extend(hits[:3])
                cuisine_hits.extend(hits[:3])
                reasons.append(f"country_keywords[{country_code}]:{'/'.join(hits[:3])}=>{' > '.join([p for p in segment if p])}")

        for rule in self.compiled_negative_keyword_rules:
            hits = self._match_rule_hits(haystack, fuzzy_tokens, rule, False)
            if hits:
                family = tuple(rule['family'])
                family_scores[family] = family_scores.get(family, 0.0) + float(rule.get('weight', 0.0))
                reasons.append(f"negative_keywords:{'/'.join(hits[:3])}=>{family[0]}>{family[1]}")

        price_bucket = self._extract_price_bucket(row.get('price_range_simplified', ''), row.get('customer_reported_price_range', ''))
        if price_bucket in PRICE_SCORE_RULES:
            for family, delta, reason in PRICE_SCORE_RULES[price_bucket]:
                family_scores[family] = family_scores.get(family, 0.0) + delta
                reasons.append(reason)

        if not detailed_scores:
            return self._empty_result(langs, country_code, price_bucket, family_scores, options.emit_debug_columns)

        best_family = max(family_scores.items(), key=lambda item: item[1])[0]
        family_candidates = [(path, score) for path, score in detailed_scores.items() if path[:2] == best_family] or list(detailed_scores.items())
        best_path, best_score = max(family_candidates, key=lambda item: item[1])
        confidence = self._compute_confidence(best_score, family_scores, has_main_type=bool(main_type_path))
        result = {
            'fyre_market_segment_type0': best_path[0],
            'fyre_market_segment_type1': best_path[1],
            'fyre_market_segment_type2': best_path[2],
            'fyre_market_segment_type3': best_path[3],
            'segmentation_confidence': confidence,
            'segmentation_reasons': ' || '.join(reasons[:16]),
            'base_main_type_path': ' > '.join([p for p in main_type_path if p]),
            'all_types_paths_considered': ' || '.join(all_types_paths[:12]),
            'keyword_hits': ' | '.join(keyword_hits[:16]),
            'language_scope': ','.join(langs),
        }
        if options.emit_debug_columns:
            result.update({
                'resolved_country_code': country_code,
                'price_signal_bucket': price_bucket,
                'family_score_snapshot': json.dumps({f'{k[0]}>{k[1]}': round(v, 2) for k, v in family_scores.items()}, ensure_ascii=False),
                'cuisine_keyword_hits': ' | '.join(cuisine_hits[:16]),
            })
        return result

    def _resolve_country_code(self, row: pd.Series, options: MarketSegmenterOptions) -> str:
        raw_code = _normalize_text(row.get('country_code', '') or options.country_default or '')
        if raw_code:
            if len(raw_code) == 2:
                return raw_code.upper()
            return COUNTRY_NAME_TO_CODE.get(raw_code, raw_code[:2].upper())
        raw_country = _normalize_text(row.get('country', '') or options.country_default or '')
        return COUNTRY_NAME_TO_CODE.get(raw_country, raw_country[:2].upper() if raw_country else '')

    def _extract_price_bucket(self, simplified_value: str, reported_value: str) -> str:
        values = [_normalize_text(simplified_value), _normalize_text(reported_value)]
        for value in values:
            if not value:
                continue
            for bucket, aliases in PRICE_BUCKETS.items():
                if value in aliases or any(alias in value for alias in aliases):
                    return bucket
            if '$$$$' in value:
                return 'very_high'
            if '$$$' in value:
                return 'high'
            if '$$' in value:
                return 'mid'
            if '$' in value:
                return 'low'
        return ''

    def _match_rule_hits(self, haystack: str, fuzzy_tokens: tuple[str, ...], rule: dict, typo_tolerance: bool) -> list[str]:
        keywords = rule.get('normalized_keywords') or tuple(_normalize_text(kw) for kw in rule.get('keywords', []))
        return [kw for kw in keywords if self._contains_keyword(haystack, fuzzy_tokens, kw, typo_tolerance)]

    def _compute_confidence(self, best_score: float, family_scores: dict[tuple[str, str], float], has_main_type: bool) -> float:
        ordered = sorted(family_scores.values(), reverse=True)
        margin = ordered[0] - ordered[1] if len(ordered) > 1 else ordered[0]
        base = 0.45 if has_main_type else 0.28
        confidence = base + min(0.28, best_score / 260.0) + min(0.18, max(margin, 0.0) / 30.0)
        return round(min(0.99, max(0.15, confidence)), 4)

    def _empty_result(self, langs: list[str], country_code: str, price_bucket: str, family_scores: dict[tuple[str, str], float], emit_debug_columns: bool) -> dict:
        result = {
            'fyre_market_segment_type0': '',
            'fyre_market_segment_type1': '',
            'fyre_market_segment_type2': '',
            'fyre_market_segment_type3': '',
            'segmentation_confidence': 0.15,
            'segmentation_reasons': 'no_rule_match',
            'base_main_type_path': '',
            'all_types_paths_considered': '',
            'keyword_hits': '',
            'language_scope': ','.join(langs),
        }
        if emit_debug_columns:
            result.update({
                'resolved_country_code': country_code,
                'price_signal_bucket': price_bucket,
                'family_score_snapshot': json.dumps({f'{k[0]}>{k[1]}': round(v, 2) for k, v in family_scores.items()}, ensure_ascii=False),
                'cuisine_keyword_hits': '',
            })
        return result

    def _map_google_type(self, value: str):
        normalized = _normalize_text(value)
        if not normalized:
            return ('', '', '', ''), ''
        exact = self.type_mapping.get(normalized)
        if exact is not None:
            return (
                str(exact.get('marketsegment0', '') or ''),
                str(exact.get('marketsegment1', '') or ''),
                str(exact.get('marketsegment2', '') or ''),
                str(exact.get('marketsegment3', '') or ''),
            ), str(exact.get('mapping_reason', '') or '')
        for key, row in self.type_mapping.items():
            if key and (key in normalized or normalized in key):
                return (
                    str(row.get('marketsegment0', '') or ''),
                    str(row.get('marketsegment1', '') or ''),
                    str(row.get('marketsegment2', '') or ''),
                    str(row.get('marketsegment3', '') or ''),
                ), str(row.get('mapping_reason', '') or '')
        return ('', '', '', ''), ''

    def _contains_keyword(self, haystack: str, tokens: list[str], keyword: str, typo_tolerance: bool) -> bool:
        kw = _normalize_text(keyword)
        if not kw:
            return False
        if kw in haystack:
            return True
        kwt = kw.split()
        if len(kwt) == 1:
            return any(
                token == kw or (typo_tolerance and len(kw) >= 5 and difflib.SequenceMatcher(None, token, kw).ratio() >= 0.86)
                for token in tokens
            )
        for start in range(0, max(len(tokens) - len(kwt) + 1, 0)):
            phrase = ' '.join(tokens[start:start + len(kwt)])
            if phrase == kw:
                return True
            if typo_tolerance and difflib.SequenceMatcher(None, phrase, kw).ratio() >= 0.88:
                return True
        return False
