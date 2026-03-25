from __future__ import annotations

import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def normalize_text(value: object) -> str:
    return '' if value is None else str(value).strip()


def detect_csv_delimiter(path: Path) -> str:
    sample = path.read_text(encoding='utf-8-sig', errors='ignore')[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        return dialect.delimiter
    except Exception:
        return ';' if sample.count(';') > sample.count(',') else ','


def iter_rows(path: Path, sheet_name: str | None = None):
    suffix = path.suffix.lower()
    if suffix in {'.csv', '.txt'}:
        delimiter = detect_csv_delimiter(path)
        with path.open('r', encoding='utf-8-sig', newline='') as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            headers = list(reader.fieldnames or [])
            for row in reader:
                yield headers, {str(k): '' if v is None else str(v) for k, v in row.items()}
        return
    if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            wb.close()
            return
        headers = ['' if v is None else str(v) for v in raw_headers]
        for values in rows:
            row = {}
            for idx, header in enumerate(headers):
                value = values[idx] if idx < len(values) else ''
                row[header] = '' if value is None else str(value)
            yield headers, row
        wb.close()
        return
    raise ValueError(f'Format non supporté pour AI review: {path.suffix}')


def estimate_total_rows(path: Path, sheet_name: str | None = None) -> int | None:
    suffix = path.suffix.lower()
    try:
        if suffix in {'.csv', '.txt'}:
            with path.open('r', encoding='utf-8-sig', errors='ignore') as fh:
                total = sum(1 for _ in fh)
            return max(total - 1, 0)
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            total = max((ws.max_row or 1) - 1, 0)
            wb.close()
            return total
    except Exception:
        return None
    return None


def split_multi_value(raw: str) -> list[str]:
    text = normalize_text(raw)
    if not text:
        return []
    if text.startswith('[') and text.endswith(']'):
        try:
            data = json.loads(text)
            if isinstance(data, list):
                return [normalize_text(x) for x in data if normalize_text(x)]
        except Exception:
            pass
    return [part.strip() for part in re.split(r'[\n\r|;,]+', text) if part.strip()]


def csv_safe(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value).replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')


def first_non_empty(row: dict[str, str], candidates: Iterable[str]) -> str:
    for key in candidates:
        value = normalize_text(row.get(key, ''))
        if value:
            return value
    return ''


def normalized_text(value: object) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ''
    text = ''.join(ch for ch in unicodedata.normalize('NFKD', text) if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', text)
