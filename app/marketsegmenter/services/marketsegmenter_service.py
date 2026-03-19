from __future__ import annotations

import csv
import difflib
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook

ProgressCallback = Callable[[int, str], None]
LogCallback = Callable[[str], None]

MARKETSEGMENTER_MAPPING_FIELDS = [
    'place_id', 'name', 'country', 'country_code', 'city', 'postal_code',
    'main_type', 'all_types', 'description_1', 'description_2', 'description_3',
    'price_range_simplified', 'customer_reported_price_range', 'reviews_tags',
    'characteristics', 'hotel_additional_informations', 'website_title',
    'website_meta_description'
]
MARKETSEGMENTER_REQUIRED_FIELDS = {'name', 'main_type'}

MARKETSEGMENTER_COLUMN_ALIASES = {
    'place_id': ['Place Id', 'place_id', 'id', 'CID', 'MID'],
    'name': ['Name', 'name'],
    'country': ['Country', 'country'],
    'country_code': ['Country code', 'country_code', 'country iso', 'country_iso'],
    'city': ['City', 'city', 'locality'],
    'postal_code': ['Postal code', 'postal_code', 'postcode', 'zip'],
    'main_type': ['Main type', 'main_type'],
    'all_types': ['All types', 'all_types'],
    'description_1': ['Description 1', 'description_1'],
    'description_2': ['Description 2', 'description_2'],
    'description_3': ['Description 3', 'description_3'],
    'price_range_simplified': ['Price range simplified', 'price_range_simplified'],
    'customer_reported_price_range': [
        'Customers reported price range', 'Customer reported price range',
        'customer_reported_price_range'
    ],
    'reviews_tags': ['Reviews tags', 'reviews_tags'],
    'characteristics': ['Characteristics', 'characteristics'],
    'hotel_additional_informations': [
        'Hotel additional informations', 'hotel_additional_informations',
        'Hotel additional information', 'hotel_additional_information'
    ],
    'website_title': ['Website title', 'website_title'],
    'website_meta_description': ['Website meta description', 'website_meta_description'],
}

COUNTRY_LANGS = {
    'FR': ['fr', 'en'], 'BE': ['fr', 'nl', 'en'], 'CH': ['fr', 'de', 'it', 'en'],
    'ES': ['es', 'en'], 'PT': ['pt', 'en'], 'IT': ['it', 'en'], 'DE': ['de', 'en'],
    'AT': ['de', 'en'], 'NL': ['nl', 'en'], 'GB': ['en'], 'IE': ['en'],
    'US': ['en'], 'CA': ['en', 'fr'], 'MX': ['es', 'en'], 'BR': ['pt', 'en'],
    'IL': ['he', 'en'], 'GR': ['el', 'en'], 'TR': ['tr', 'en'],
}
COUNTRY_NAME_TO_CODE = {
    'france': 'FR', 'french republic': 'FR', 'belgium': 'BE', 'belgique': 'BE',
    'switzerland': 'CH', 'suisse': 'CH', 'spain': 'ES', 'espana': 'ES', 'espagne': 'ES',
    'portugal': 'PT', 'italy': 'IT', 'italia': 'IT', 'germany': 'DE', 'deutschland': 'DE',
    'austria': 'AT', 'netherlands': 'NL', 'nederland': 'NL', 'united kingdom': 'GB',
    'uk': 'GB', 'great britain': 'GB', 'england': 'GB', 'ireland': 'IE',
    'united states': 'US', 'usa': 'US', 'canada': 'CA', 'mexico': 'MX', 'brazil': 'BR',
    'israel': 'IL', 'greece': 'GR', 'turkey': 'TR',
}

ALL_TEXT_FIELDS = [
    'name', 'main_type', 'all_types', 'description_1', 'description_2', 'description_3',
    'reviews_tags', 'characteristics', 'hotel_additional_informations',
    'website_title', 'website_meta_description'
]
DEBUG_OUTPUT_COLUMNS = [
    'resolved_country_code', 'price_signal_bucket', 'family_score_snapshot',
    'cuisine_keyword_hits'
]

BASE_FAMILY_SCORE_RULES = {
    ('horeca', 'table_service'): 0.0,
    ('horeca', 'fast_food'): 0.0,
    ('horeca', 'cafes_bars_discotheques'): 0.0,
    ('horeca', 'hotel_lodging'): 0.0,
    ('asc', 'food_store'): 0.0,
    ('asc', 'other_types_of_businesses'): 0.0,
    ('leisure', 'on_site_catering_events'): 0.0,
}

PRICE_BUCKETS = {
    'very_low': {'$', 'very low', 'budget', 'cheap', 'inexpensive', 'bon marche', 'pas cher', 'economique', 'economy', 'low cost', '1'},
    'low': {'low', 'moderate-low', 'low to moderate', 'moderate low', '2'},
    'mid': {'medium', 'moderate', 'mid', 'average', 'normal', '3'},
    'high': {'high', 'expensive', 'upscale', 'premium', '4'},
    'very_high': {'very high', 'luxury', 'fine dining', '5'},
}
PRICE_SCORE_RULES = {
    'very_low': [
        (('horeca', 'fast_food'), 7.0, 'price:very_low=>fast_food'),
        (('horeca', 'table_service'), -2.0, 'price:very_low=>table_service_malus'),
        (('horeca', 'cafes_bars_discotheques'), 1.5, 'price:very_low=>bars_minor'),
    ],
    'low': [
        (('horeca', 'fast_food'), 5.0, 'price:low=>fast_food'),
        (('horeca', 'table_service'), -1.0, 'price:low=>table_service_malus'),
    ],
    'mid': [
        (('horeca', 'table_service'), 2.0, 'price:mid=>table_service'),
        (('horeca', 'cafes_bars_discotheques'), 1.0, 'price:mid=>bars'),
    ],
    'high': [
        (('horeca', 'table_service'), 5.0, 'price:high=>table_service'),
        (('horeca', 'fast_food'), -2.5, 'price:high=>fast_food_malus'),
        (('horeca', 'cafes_bars_discotheques'), 2.0, 'price:high=>bars'),
    ],
    'very_high': [
        (('horeca', 'table_service'), 6.0, 'price:very_high=>table_service'),
        (('horeca', 'fast_food'), -3.0, 'price:very_high=>fast_food_malus'),
        (('horeca', 'cafes_bars_discotheques'), 2.5, 'price:very_high=>bars'),
    ],
}

KEYWORD_RULES = [
    {'segment': ['horeca', 'fast_food', 'fast_food_burgers_chicken', ''], 'family': ('horeca', 'fast_food'), 'weight': 14,
     'keywords': ['burger', 'burgers', 'hamburger', 'cheeseburger', 'fried chicken', 'crispy chicken', 'wings', 'nuggets', 'chicken burger', 'hamburgr', 'burguer', 'berger', 'bruger']},
    {'segment': ['horeca', 'fast_food', 'fast_food_kebab_tacos_mexican', ''], 'family': ('horeca', 'fast_food'), 'weight': 14,
     'keywords': ['kebab', 'kebap', 'shawarma', 'shawerma', 'tacos', 'taco', 'burrito', 'quesadilla', 'doner', 'dürüm', 'durum', 'gyro', 'gyros']},
    {'segment': ['horeca', 'fast_food', 'fast_food_sandwiches_bagels_salads', ''], 'family': ('horeca', 'fast_food'), 'weight': 12,
     'keywords': ['sandwich', 'sandwiche', 'bagel', 'bagels', 'salad bar', 'salads', 'wrap', 'panini', 'bocadillo', 'sub', 'subs', 'hoagie', 'poke bowl', 'poké', 'bowl', 'boul']},
    {'segment': ['horeca', 'fast_food', 'pizza_pasta', ''], 'family': ('horeca', 'fast_food'), 'weight': 12,
     'keywords': ['pizza', 'pizzas', 'slice', 'slices', 'pasta', 'lasagna', 'spaghetti', 'tagliatelle', 'pizzeria', 'pizzaria', 'pizzaa']},
    {'segment': ['horeca', 'fast_food', 'asian', ''], 'family': ('horeca', 'fast_food'), 'weight': 11,
     'keywords': ['ramen', 'udon', 'wok', 'noodle', 'noodles', 'yakisoba', 'sushi burrito', 'bao', 'gyoza', 'bento', 'bento box']},
    {'segment': ['horeca', 'fast_food', 'coffee_shops_ice_cream_parlors_kiosks', ''], 'family': ('horeca', 'fast_food'), 'weight': 12,
     'keywords': ['coffee', 'cafe', 'café', 'espresso', 'latte', 'cappuccino', 'barista', 'tea room', 'ice cream', 'gelato', 'frozen yogurt', 'froyo', 'smoothie', 'juice bar', 'kiosk', 'kiosque', 'cafee', 'cofee']},
    {'segment': ['horeca', 'fast_food', 'self_service_cafeterias', ''], 'family': ('horeca', 'fast_food'), 'weight': 11,
     'keywords': ['self service', 'self-service', 'cafeteria', 'canteen', 'buffet line', 'tray service', 'counter service', 'service au comptoir']},
    {'segment': ['horeca', 'fast_food', 'other_fast_food', ''], 'family': ('horeca', 'fast_food'), 'weight': 9,
     'keywords': ['takeaway', 'take away', 'take-out', 'takeout', 'delivery only', 'quick bite', 'grab and go', 'grab&go', 'street food', 'snack bar', 'snack', 'food truck', 'drive thru', 'drive-thru', 'drive in']},

    {'segment': ['horeca', 'table_service', 'fine_dining', ''], 'family': ('horeca', 'table_service'), 'weight': 16,
     'keywords': ['fine dining', 'chef tasting', 'degustation', 'menu degustation', 'michelin', 'sommelier', 'gourmet', 'haute cuisine', 'signature menu', 'reservation recommended', 'white tablecloth']},
    {'segment': ['horeca', 'table_service', 'brasseries_cafe_restaurants', ''], 'family': ('horeca', 'table_service'), 'weight': 10,
     'keywords': ['brasserie', 'bistrot', 'bistro', 'cafe restaurant', 'café restaurant', 'trattoria', 'osteria', 'pub food', 'family restaurant', 'diner', 'sit-down', 'table service', 'service en salle']},
    {'segment': ['horeca', 'table_service', 'creperies', ''], 'family': ('horeca', 'table_service'), 'weight': 13,
     'keywords': ['crepe', 'crêpe', 'creperie', 'crêperie', 'galette', 'galettes']},
    {'segment': ['horeca', 'table_service', 'themed_dining_asian', ''], 'family': ('horeca', 'table_service'), 'weight': 14,
     'keywords': ['sushi', 'sashimi', 'ramen restaurant', 'thai curry', 'pho', 'dim sum', 'dumpling', 'izakaya', 'teppanyaki', 'bibimbap', 'kimchi', 'pad thai', 'bao buns', 'wonton', 'yakitori']},
    {'segment': ['horeca', 'table_service', 'themed_dining_fish_seafood', ''], 'family': ('horeca', 'table_service'), 'weight': 14,
     'keywords': ['seafood', 'fish', 'oyster', 'oysters', 'lobster', 'crab', 'shrimp', 'prawns', 'moules', 'fruits de mer', 'poisson', 'ceviche']},
    {'segment': ['horeca', 'table_service', 'themed_dining_grills_meat_specialties', ''], 'family': ('horeca', 'table_service'), 'weight': 14,
     'keywords': ['steak', 'steakhouse', 'bbq', 'barbecue', 'grill', 'grillades', 'ribs', 'smoked meat', 'rotisserie', 'churrasco', 'asado', 'meat lovers', 'braai']},
    {'segment': ['horeca', 'table_service', 'themed_dining_italian_pizzerias', ''], 'family': ('horeca', 'table_service'), 'weight': 14,
     'keywords': ['italian', 'italiano', 'ristorante', 'trattoria', 'pizzeria', 'pizza napolitaine', 'napoletana', 'gnocchi', 'risotto', 'antipasti', 'pasta fresca', 'pizza oven']},
    {'segment': ['horeca', 'table_service', 'themed_dining_other_specialties', ''], 'family': ('horeca', 'table_service'), 'weight': 13,
     'keywords': ['lebanese', 'libanais', 'shawarma plate', 'mezze', 'mezzeh', 'hummus', 'falafel', 'turkish', 'greek', 'mexican restaurant', 'indian', 'curry house', 'ethiopian', 'peruvian', 'argentinian', 'moroccan', 'tagine', 'couscous', 'georgian', 'armenian', 'israeli', 'mediterranean restaurant', 'halal grill', 'syrian', 'afghani']},
    {'segment': ['horeca', 'table_service', 'traditional_dining', ''], 'family': ('horeca', 'table_service'), 'weight': 9,
     'keywords': ['restaurant', 'restaurante', 'ristorante', 'resto', 'family meals', 'dining room', 'plat du jour', 'menu du jour', 'sit down meal', 'reservation', 'terrace dining']},

    {'segment': ['horeca', 'cafes_bars_discotheques', 'beer_bars_beer_temples_pubs', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 13,
     'keywords': ['pub', 'ale', 'beer', 'brewery', 'brewpub', 'taproom', 'craft beer', 'irish pub', 'sports pub', 'bier', 'cerveza artesanal']},
    {'segment': ['horeca', 'cafes_bars_discotheques', 'cocktail_bars', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 13,
     'keywords': ['cocktail', 'mixology', 'speakeasy', 'martini', 'negroni', 'mojito', 'bartender', 'signature cocktails']},
    {'segment': ['horeca', 'cafes_bars_discotheques', 'wine_tapas_bars', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 13,
     'keywords': ['wine bar', 'wine list', 'wines by the glass', 'tapas', 'pinchos', 'enoteca', 'vinoteca', 'charcuterie board', 'apero']},
    {'segment': ['horeca', 'cafes_bars_discotheques', 'clubs_nightclubs_party_bars', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 13,
     'keywords': ['nightclub', 'night club', 'dj set', 'dance floor', 'vip table', 'party bar', 'karaoke', 'late night', 'discotheque']},
    {'segment': ['horeca', 'cafes_bars_discotheques', 'traditional_bars', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 10,
     'keywords': ['bar', 'bar tabac', 'tabac', 'café-bar', 'local bar', 'bar lounge']},

    {'segment': ['horeca', 'hotel_lodging', 'hotels', ''], 'family': ('horeca', 'hotel_lodging'), 'weight': 14,
     'keywords': ['hotel', 'resort', 'spa hotel', 'boutique hotel', 'rooms available', 'check-in', 'concierge', 'room service']},
    {'segment': ['horeca', 'hotel_lodging', 'campsites', ''], 'family': ('horeca', 'hotel_lodging'), 'weight': 13,
     'keywords': ['camping', 'campground', 'caravan', 'rv park', 'glamping', 'pitch', 'bungalow park']},
    {'segment': ['horeca', 'hotel_lodging', 'other_accommodations', ''], 'family': ('horeca', 'hotel_lodging'), 'weight': 11,
     'keywords': ['hostel', 'guest house', 'bed and breakfast', 'bed & breakfast', 'holiday rental', 'aparthotel', 'serviced apartment']},

    {'segment': ['asc', 'food_store', 'bakeries_pastry_shops', ''], 'family': ('asc', 'food_store'), 'weight': 13,
     'keywords': ['bakery', 'boulangerie', 'patisserie', 'pastry', 'viennoiserie', 'croissant', 'bread', 'artisan bread', 'gateaux']},
    {'segment': ['asc', 'food_store', 'butchers_delicatessens_caterers', ''], 'family': ('asc', 'food_store'), 'weight': 13,
     'keywords': ['butcher', 'butchery', 'charcuterie', 'delicatessen', 'traiteur', 'cold cuts', 'meat shop', 'hams', 'salumi']},
    {'segment': ['asc', 'food_store', 'wine_shops', ''], 'family': ('asc', 'food_store'), 'weight': 12,
     'keywords': ['wine shop', 'caviste', 'cave a vins', 'cave à vins', 'liquor store', 'spirits', 'whisky shop', 'vinoteca']},
    {'segment': ['asc', 'food_store', 'convenience_stores', ''], 'family': ('asc', 'food_store'), 'weight': 11,
     'keywords': ['convenience store', 'mini market', 'minimarket', 'corner shop', '24/7', 'open 24 hours', 'grocery']},
    {'segment': ['asc', 'food_store', 'other_food_stores', 'tea_shop'], 'family': ('asc', 'food_store'), 'weight': 10,
     'keywords': ['tea shop', 'tea room', 'thé', 'teas', 'infusions']},

    {'segment': ['leisure', 'on_site_catering_events', 'cultural_sites', ''], 'family': ('leisure', 'on_site_catering_events'), 'weight': 12,
     'keywords': ['museum', 'gallery', 'heritage', 'cultural center', 'exhibition']},
    {'segment': ['leisure', 'on_site_catering_events', 'leisure_sites', 'beach_club'], 'family': ('leisure', 'on_site_catering_events'), 'weight': 12,
     'keywords': ['beach club', 'sunbeds', 'pool club']},
    {'segment': ['leisure', 'on_site_catering_events', 'leisure_sites', 'spa'], 'family': ('leisure', 'on_site_catering_events'), 'weight': 12,
     'keywords': ['spa', 'wellness', 'massage', 'hammam', 'sauna']},
    {'segment': ['leisure', 'on_site_catering_events', 'leisure_sites', 'amusement_recreation'], 'family': ('leisure', 'on_site_catering_events'), 'weight': 12,
     'keywords': ['amusement park', 'theme park', 'water park', 'arcade', 'roller coaster']},
    {'segment': ['leisure', 'on_site_catering_events', 'sports_sites', ''], 'family': ('leisure', 'on_site_catering_events'), 'weight': 12,
     'keywords': ['stadium', 'sports complex', 'golf', 'tennis club', 'padel', 'fitness center']},
]

COUNTRY_KEYWORD_RULES = [
    {'countries': ['FR', 'BE', 'CH'], 'segment': ['horeca', 'table_service', 'creperies', ''], 'family': ('horeca', 'table_service'), 'weight': 12,
     'keywords': ['galette', 'galettes', 'crêperie', 'creperie']},
    {'countries': ['ES'], 'segment': ['horeca', 'cafes_bars_discotheques', 'wine_tapas_bars', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 12,
     'keywords': ['tapas', 'pinchos', 'cerveceria', 'cervecería', 'taberna']},
    {'countries': ['IT'], 'segment': ['horeca', 'table_service', 'themed_dining_italian_pizzerias', ''], 'family': ('horeca', 'table_service'), 'weight': 12,
     'keywords': ['ristorante', 'trattoria', 'osteria', 'pizzeria', 'pizza al taglio', 'forno a legna', 'apericena']},
    {'countries': ['DE', 'AT', 'CH'], 'segment': ['horeca', 'cafes_bars_discotheques', 'beer_bars_beer_temples_pubs', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 12,
     'keywords': ['biergarten', 'brauhaus', 'kneipe', 'bierkeller']},
    {'countries': ['GB', 'IE'], 'segment': ['horeca', 'cafes_bars_discotheques', 'beer_bars_beer_temples_pubs', ''], 'family': ('horeca', 'cafes_bars_discotheques'), 'weight': 12,
     'keywords': ['gastropub', 'public house', 'pub grub']},
    {'countries': ['IL'], 'segment': ['horeca', 'table_service', 'themed_dining_other_specialties', ''], 'family': ('horeca', 'table_service'), 'weight': 12,
     'keywords': ['falafel', 'sabich', 'shakshuka', 'hummus', 'shawarma', 'shwarma']},
]

NEGATIVE_KEYWORDS = [
    {'family': ('horeca', 'table_service'), 'weight': -6.0,
     'keywords': ['takeaway', 'take away', 'delivery only', 'food truck', 'drive thru', 'drive-thru', 'grab and go']},
    {'family': ('horeca', 'fast_food'), 'weight': -5.0,
     'keywords': ['reservation', 'tasting menu', 'sommelier', 'chef table', 'dining room', 'white tablecloth']},
]


def _noop_progress(percent: int, message: str) -> None:
    return None


def _noop_log(message: str) -> None:
    return None


def _normalize_text(value: object) -> str:
    text = '' if value is None else str(value)
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = re.sub(r"[\u2019'`´]+", ' ', text)
    text = re.sub(r'[^a-z0-9&/+ ]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _split_csv_like(raw: str) -> list[str]:
    raw = (raw or '').strip()
    return [part.strip() for part in re.split(r'[,\|;/]+', raw) if part.strip()] if raw else []


def suggest_column_mapping(columns: list[str]) -> dict[str, str]:
    normalized_to_original = {_normalize_text(col).replace(' ', '_'): col for col in columns}
    suggestions: dict[str, str] = {}
    used: set[str] = set()
    for target, aliases in MARKETSEGMENTER_COLUMN_ALIASES.items():
        for alias in aliases:
            candidate = normalized_to_original.get(_normalize_text(alias).replace(' ', '_'))
            if candidate and candidate not in used:
                suggestions[target] = candidate
                used.add(candidate)
                break
        if target in suggestions:
            continue
        for original in columns:
            if original in used:
                continue
            n = _normalize_text(original)
            if any(_normalize_text(alias) in n or n in _normalize_text(alias) for alias in aliases):
                suggestions[target] = original
                used.add(original)
                break
    return suggestions



def _detect_csv_delimiter(path: Path) -> str:
    sample = path.read_text(encoding='utf-8-sig', errors='ignore')[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;	|')
        return dialect.delimiter
    except Exception:
        return ','


def _iter_source_rows(path: Path, sheet_name: str | None = None):
    suffix = path.suffix.lower()
    if suffix in {'.csv', '.txt'}:
        delimiter = _detect_csv_delimiter(path)
        with path.open('r', encoding='utf-8-sig', newline='') as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            headers = list(reader.fieldnames or [])
            for row in reader:
                yield headers, {str(k): '' if v is None else str(v) for k, v in row.items()}
        return
    if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            wb.close()
            return
        headers = ['' if v is None else str(v) for v in raw_headers]
        for values in rows:
            row = {}
            for idx, header in enumerate(headers):
                value = values[idx] if idx < len(values) else ''
                row[header] = '' if value is None else str(value)
            yield headers, row
        wb.close()
        return
    raise ValueError(f'Format non supporté pour le market segmenter: {path.suffix}')


def _estimate_total_rows(path: Path, sheet_name: str | None = None) -> int | None:
    suffix = path.suffix.lower()
    try:
        if suffix in {'.csv', '.txt'}:
            with path.open('r', encoding='utf-8-sig', errors='ignore') as fh:
                total = sum(1 for _ in fh)
            return max(total - 1, 0)
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            total = max((ws.max_row or 1) - 1, 0)
            wb.close()
            return total
    except Exception:
        return None
    return None


def _prepare_output_headers(source_headers: list[str], mapping: dict[str, str]) -> list[str]:
    reverse = {source: target for target, source in mapping.items() if source in source_headers}
    output_headers = [reverse.get(col, col) for col in source_headers]
    existing = set(output_headers)
    for field in MARKETSEGMENTER_MAPPING_FIELDS:
        if field not in existing:
            output_headers.append(field)
            existing.add(field)
    return output_headers


def _map_row_dict(row: dict[str, str], source_headers: list[str], mapping: dict[str, str], output_headers: list[str]) -> dict[str, str]:
    reverse = {source: target for target, source in mapping.items() if source in source_headers}
    out: dict[str, str] = {}
    for col in source_headers:
        out[reverse.get(col, col)] = '' if row.get(col) is None else str(row.get(col, ''))
    for field in output_headers:
        out.setdefault(field, '')
    return out


def _csv_safe(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = str(value)
    return text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
def inspect_marketsegmenter_file(uploaded_file) -> dict:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
        xls = pd.ExcelFile(uploaded_file)
        sheets = []
        for name in xls.sheet_names[:10]:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, sheet_name=name, nrows=20)
            columns = [str(c) for c in df.columns]
            suggestions = suggest_column_mapping(columns)
            sheets.append({
                'name': name,
                'max_row': None,
                'max_column': len(columns),
                'preview': [columns] + df.head(20).fillna('').astype(str).values.tolist(),
                'detected_columns': columns,
                'mapping_suggestions': suggestions,
                'missing_required': sorted(MARKETSEGMENTER_REQUIRED_FIELDS - set(suggestions.keys())),
            })
        return {'filename': Path(uploaded_file.name).name, 'kind': 'excel', 'sheets': sheets}
    if suffix in {'.csv', '.txt'}:
        df = pd.read_csv(uploaded_file, nrows=20)
        columns = [str(c) for c in df.columns]
        suggestions = suggest_column_mapping(columns)
        return {
            'filename': Path(uploaded_file.name).name,
            'kind': 'csv',
            'sheets': [{
                'name': '__csv__',
                'max_row': None,
                'max_column': len(columns),
                'preview': [columns] + df.head(20).fillna('').astype(str).values.tolist(),
                'detected_columns': columns,
                'mapping_suggestions': suggestions,
                'missing_required': sorted(MARKETSEGMENTER_REQUIRED_FIELDS - set(suggestions.keys())),
            }],
        }
    raise ValueError('Inspection market segmenter disponible uniquement pour CSV et Excel.')


@dataclass
class MarketSegmenterOptions:
    marketsegmenter_sheet_name: str | None = None
    marketsegmenter_mapping: dict[str, str] = field(default_factory=dict)
    country_default: str = ''
    enable_typo_tolerance: bool = True
    emit_debug_columns: bool = True


class MarketSegmenterService:
    def __init__(self, progress_callback: ProgressCallback | None = None, log_callback: LogCallback | None = None):
        self.progress_callback = progress_callback or _noop_progress
        self.log_callback = log_callback or _noop_log
        self.type_mapping_df = self._load_type_mapping_df()
        self.type_mapping = {
            _normalize_text(row['type']): row
            for _, row in self.type_mapping_df.fillna('').iterrows()
            if str(row.get('type', '')).strip()
        }
        self.compiled_keyword_rules = self._compile_keyword_rules(KEYWORD_RULES)
        self.compiled_country_keyword_rules = self._compile_keyword_rules(COUNTRY_KEYWORD_RULES)
        self.compiled_negative_keyword_rules = self._compile_keyword_rules(NEGATIVE_KEYWORDS)

    def progress(self, percent: int, message: str) -> None:
        self.progress_callback(percent, message)

    def log(self, message: str) -> None:
        self.log_callback(message)

    def run(self, input_path: Path, output_path: Path, options: MarketSegmenterOptions) -> Path:
        self.progress(5, 'Chargement du fichier source du market segmenter')
        output_path.parent.mkdir(parents=True, exist_ok=True)

        debug_cols = DEBUG_OUTPUT_COLUMNS if options.emit_debug_columns else []
        stats = defaultdict(int)
        total = _estimate_total_rows(input_path, options.marketsegmenter_sheet_name) or 0
        processed = 0
        source_headers: list[str] | None = None
        projected_headers: list[str] | None = None

        with output_path.open('w', newline='', encoding='utf-8-sig') as fh:
            writer = None
            for raw_headers, raw_row in _iter_source_rows(input_path, options.marketsegmenter_sheet_name):
                if source_headers is None:
                    source_headers = list(raw_headers)
                    projected_headers = _prepare_output_headers(source_headers, options.marketsegmenter_mapping)
                    output_columns = list(projected_headers) + [
                        'fyre_market_segment_type0', 'fyre_market_segment_type1',
                        'fyre_market_segment_type2', 'fyre_market_segment_type3',
                        'segmentation_confidence', 'segmentation_reasons',
                        'base_main_type_path', 'all_types_paths_considered',
                        'keyword_hits', 'language_scope',
                    ] + debug_cols
                    writer = csv.DictWriter(
                        fh,
                        fieldnames=output_columns,
                        delimiter=';',
                        quotechar='"',
                        quoting=csv.QUOTE_ALL,
                        lineterminator='\n',
                    )
                    writer.writeheader()

                row = _map_row_dict(raw_row, source_headers or [], options.marketsegmenter_mapping, projected_headers or [])
                classified = self._classify_row(row, options)
                out = {col: _csv_safe(row.get(col, '')) for col in (projected_headers or [])}
                out.update({k: _csv_safe(v) for k, v in classified.items()})
                writer.writerow(out)
                processed += 1
                stats[classified.get('fyre_market_segment_type0') or 'unknown'] += 1
                if processed == 1 or processed % 1000 == 0 or (total and processed == total):
                    self.progress(min(96, 15 + int((processed / max(total, processed)) * 80)), f'Segmentation en cours : {processed}/{total or "?"}')

        summary = {
            'rows': int(processed),
            'macro_stats': dict(stats),
            'type_mapping_coverage': int(self.type_mapping_df['marketsegment0'].astype(str).str.strip().ne('').sum()),
            'typo_tolerance': bool(options.enable_typo_tolerance),
            'country_default': options.country_default or '',
            'country_scope_enabled': True,
            'price_signal_enabled': True,
            'culinary_keywords_enabled': True,
            'streaming_mode_enabled': True,
            'safe_csv_export': True,
        }
        output_path.with_name(output_path.stem + '_summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
        self.progress(100, 'Market segmenter terminé')
        return output_path

    def _load_type_mapping_df(self) -> pd.DataFrame:
        mapping_path = Path(__file__).resolve().parents[1] / 'data' / 'google_type_mapping_proposed.csv'
        if mapping_path.exists():
            return pd.read_csv(mapping_path).fillna('')
        return pd.DataFrame(columns=['type', 'marketsegment0', 'marketsegment1', 'marketsegment2', 'marketsegment3', 'mapping_reason'])

    def _apply_mapping(self, df: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
        reverse = {source: target for target, source in mapping.items() if source in df.columns}
        df = df.rename(columns=reverse)
        for field in MARKETSEGMENTER_MAPPING_FIELDS:
            if field not in df.columns:
                df[field] = ''
            df[field] = df[field].fillna('').astype(str)
        return df

    def _compile_keyword_rules(self, rules: list[dict]) -> list[dict]:
        compiled = []
        for rule in rules:
            normalized_keywords = []
            for kw in rule.get('keywords', []):
                nkw = _normalize_text(kw)
                if nkw:
                    normalized_keywords.append(nkw)
            compiled.append({**rule, 'normalized_keywords': tuple(dict.fromkeys(normalized_keywords))})
        return compiled

    def _classify_row(self, row: pd.Series, options: MarketSegmenterOptions) -> dict:
        country_code = self._resolve_country_code(row, options)
        langs = COUNTRY_LANGS.get(country_code, ['en'])
        haystack = _normalize_text(' | '.join(str(row.get(k, '')) for k in ALL_TEXT_FIELDS))
        fuzzy_haystack = _normalize_text(' | '.join(str(row.get(k, '')) for k in ['name', 'main_type', 'all_types', 'description_1', 'description_2', 'description_3']))
        tokens = haystack.split()
        fuzzy_tokens = tuple(sorted(set(fuzzy_haystack.split())))

        detailed_scores: dict[tuple[str, str, str, str], float] = defaultdict(float)
        family_scores: dict[tuple[str, str], float] = dict(BASE_FAMILY_SCORE_RULES)
        reasons: list[str] = []
        keyword_hits: list[str] = []
        cuisine_hits: list[str] = []

        main_type_path, main_reason = self._map_google_type(row.get('main_type', ''))
        if any(main_type_path):
            detailed_scores[main_type_path] += 100.0
            family_scores[(main_type_path[0], main_type_path[1])] = family_scores.get((main_type_path[0], main_type_path[1]), 0.0) + 16.0
            reasons.append(f"main_type:{row.get('main_type', '')}=>{' > '.join([p for p in main_type_path if p])}")
            if main_reason:
                reasons.append(f"main_type_reason:{main_reason}")

        all_types_paths: list[str] = []
        for idx, candidate_type in enumerate(_split_csv_like(row.get('all_types', ''))[:25], start=1):
            path, reason = self._map_google_type(candidate_type)
            if any(path):
                same_family = path[:2] == main_type_path[:2]
                detailed_scores[path] += 45.0 if same_family else 70.0
                family_scores[(path[0], path[1])] = family_scores.get((path[0], path[1]), 0.0) + (8.0 if same_family else 12.0)
                all_types_paths.append(' > '.join([p for p in path if p]))
                reasons.append(f"all_types[{idx}]:{candidate_type}=>{' > '.join([p for p in path if p])}")
                if reason:
                    reasons.append(f"all_types_reason[{idx}]:{reason}")

        ordered_pre = sorted(family_scores.values(), reverse=True)
        prelim_gap = ordered_pre[0] - ordered_pre[1] if len(ordered_pre) > 1 else (ordered_pre[0] if ordered_pre else 0.0)
        deep_scan = prelim_gap < 18.0

        for rule in self.compiled_keyword_rules:
            hits = self._match_rule_hits(haystack, fuzzy_tokens, rule, options.enable_typo_tolerance and deep_scan)
            if hits:
                segment = tuple((rule.get('segment') or ['', '', '', ''])[:4])
                detailed_scores[segment] += float(rule.get('weight', 10))
                family = tuple(rule.get('family') or segment[:2])
                family_scores[family] = family_scores.get(family, 0.0) + float(rule.get('weight', 10)) * 0.65
                keyword_hits.extend(hits[:3])
                cuisine_hits.extend(hits[:3])
                reasons.append(f"keywords:{'/'.join(hits[:3])}=>{' > '.join([p for p in segment if p])}")

        for rule in self.compiled_country_keyword_rules:
            if country_code and country_code not in set(rule.get('countries', [])):
                continue
            hits = self._match_rule_hits(haystack, fuzzy_tokens, rule, options.enable_typo_tolerance and deep_scan)
            if hits:
                segment = tuple((rule.get('segment') or ['', '', '', ''])[:4])
                detailed_scores[segment] += float(rule.get('weight', 10))
                family = tuple(rule.get('family') or segment[:2])
                family_scores[family] = family_scores.get(family, 0.0) + float(rule.get('weight', 10)) * 0.7
                keyword_hits.extend(hits[:3])
                cuisine_hits.extend(hits[:3])
                reasons.append(f"country_keywords[{country_code}]:{'/'.join(hits[:3])}=>{' > '.join([p for p in segment if p])}")

        for rule in self.compiled_negative_keyword_rules:
            hits = self._match_rule_hits(haystack, fuzzy_tokens, rule, False)
            if hits:
                family = tuple(rule['family'])
                family_scores[family] = family_scores.get(family, 0.0) + float(rule.get('weight', 0.0))
                reasons.append(f"negative_keywords:{'/'.join(hits[:3])}=>{family[0]}>{family[1]}")

        price_bucket = self._extract_price_bucket(row.get('price_range_simplified', ''), row.get('customer_reported_price_range', ''))
        if price_bucket in PRICE_SCORE_RULES:
            for family, delta, reason in PRICE_SCORE_RULES[price_bucket]:
                family_scores[family] = family_scores.get(family, 0.0) + delta
                reasons.append(reason)

        if not detailed_scores:
            return self._empty_result(langs, country_code, price_bucket, family_scores, options.emit_debug_columns)

        best_family = max(family_scores.items(), key=lambda item: item[1])[0]
        family_candidates = [(path, score) for path, score in detailed_scores.items() if path[:2] == best_family] or list(detailed_scores.items())
        best_path, best_score = max(family_candidates, key=lambda item: item[1])
        confidence = self._compute_confidence(best_score, family_scores, has_main_type=bool(main_type_path))
        result = {
            'fyre_market_segment_type0': best_path[0],
            'fyre_market_segment_type1': best_path[1],
            'fyre_market_segment_type2': best_path[2],
            'fyre_market_segment_type3': best_path[3],
            'segmentation_confidence': confidence,
            'segmentation_reasons': ' || '.join(reasons[:16]),
            'base_main_type_path': ' > '.join([p for p in main_type_path if p]),
            'all_types_paths_considered': ' || '.join(all_types_paths[:12]),
            'keyword_hits': ' | '.join(keyword_hits[:16]),
            'language_scope': ','.join(langs),
        }
        if options.emit_debug_columns:
            result.update({
                'resolved_country_code': country_code,
                'price_signal_bucket': price_bucket,
                'family_score_snapshot': json.dumps({f'{k[0]}>{k[1]}': round(v, 2) for k, v in family_scores.items()}, ensure_ascii=False),
                'cuisine_keyword_hits': ' | '.join(cuisine_hits[:16]),
            })
        return result

    def _resolve_country_code(self, row: pd.Series, options: MarketSegmenterOptions) -> str:
        raw_code = _normalize_text(row.get('country_code', '') or options.country_default or '')
        if raw_code:
            if len(raw_code) == 2:
                return raw_code.upper()
            return COUNTRY_NAME_TO_CODE.get(raw_code, raw_code[:2].upper())
        raw_country = _normalize_text(row.get('country', '') or options.country_default or '')
        return COUNTRY_NAME_TO_CODE.get(raw_country, raw_country[:2].upper() if raw_country else '')

    def _extract_price_bucket(self, simplified_value: str, reported_value: str) -> str:
        values = [_normalize_text(simplified_value), _normalize_text(reported_value)]
        for value in values:
            if not value:
                continue
            for bucket, aliases in PRICE_BUCKETS.items():
                if value in aliases or any(alias in value for alias in aliases):
                    return bucket
            if '$$$$' in value:
                return 'very_high'
            if '$$$' in value:
                return 'high'
            if '$$' in value:
                return 'mid'
            if '$' in value:
                return 'low'
        return ''

    def _match_rule_hits(self, haystack: str, fuzzy_tokens: tuple[str, ...], rule: dict, typo_tolerance: bool) -> list[str]:
        keywords = rule.get('normalized_keywords') or tuple(_normalize_text(kw) for kw in rule.get('keywords', []))
        return [kw for kw in keywords if self._contains_keyword(haystack, fuzzy_tokens, kw, typo_tolerance)]

    def _compute_confidence(self, best_score: float, family_scores: dict[tuple[str, str], float], has_main_type: bool) -> float:
        ordered = sorted(family_scores.values(), reverse=True)
        margin = ordered[0] - ordered[1] if len(ordered) > 1 else ordered[0]
        base = 0.45 if has_main_type else 0.28
        confidence = base + min(0.28, best_score / 260.0) + min(0.18, max(margin, 0.0) / 30.0)
        return round(min(0.99, max(0.15, confidence)), 4)

    def _empty_result(self, langs: list[str], country_code: str, price_bucket: str, family_scores: dict[tuple[str, str], float], emit_debug_columns: bool) -> dict:
        result = {
            'fyre_market_segment_type0': '',
            'fyre_market_segment_type1': '',
            'fyre_market_segment_type2': '',
            'fyre_market_segment_type3': '',
            'segmentation_confidence': 0.15,
            'segmentation_reasons': 'no_rule_match',
            'base_main_type_path': '',
            'all_types_paths_considered': '',
            'keyword_hits': '',
            'language_scope': ','.join(langs),
        }
        if emit_debug_columns:
            result.update({
                'resolved_country_code': country_code,
                'price_signal_bucket': price_bucket,
                'family_score_snapshot': json.dumps({f'{k[0]}>{k[1]}': round(v, 2) for k, v in family_scores.items()}, ensure_ascii=False),
                'cuisine_keyword_hits': '',
            })
        return result

    def _map_google_type(self, value: str):
        normalized = _normalize_text(value)
        if not normalized:
            return ('', '', '', ''), ''
        exact = self.type_mapping.get(normalized)
        if exact is not None:
            return (
                str(exact.get('marketsegment0', '') or ''),
                str(exact.get('marketsegment1', '') or ''),
                str(exact.get('marketsegment2', '') or ''),
                str(exact.get('marketsegment3', '') or ''),
            ), str(exact.get('mapping_reason', '') or '')
        for key, row in self.type_mapping.items():
            if key and (key in normalized or normalized in key):
                return (
                    str(row.get('marketsegment0', '') or ''),
                    str(row.get('marketsegment1', '') or ''),
                    str(row.get('marketsegment2', '') or ''),
                    str(row.get('marketsegment3', '') or ''),
                ), str(row.get('mapping_reason', '') or '')
        return ('', '', '', ''), ''

    def _contains_keyword(self, haystack: str, tokens: list[str], keyword: str, typo_tolerance: bool) -> bool:
        kw = _normalize_text(keyword)
        if not kw:
            return False
        if kw in haystack:
            return True
        kwt = kw.split()
        if len(kwt) == 1:
            return any(
                token == kw or (typo_tolerance and len(kw) >= 5 and difflib.SequenceMatcher(None, token, kw).ratio() >= 0.86)
                for token in tokens
            )
        for start in range(0, max(len(tokens) - len(kwt) + 1, 0)):
            phrase = ' '.join(tokens[start:start + len(kwt)])
            if phrase == kw:
                return True
            if typo_tolerance and difflib.SequenceMatcher(None, phrase, kw).ratio() >= 0.88:
                return True
        return False
