from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook
from slugify import slugify


ProgressCallback = Callable[[int, str], None]
LogCallback = Callable[[str], None]


CHAINES_CSV_PATH = Path(__file__).resolve().parent.parent.parent / 'legacy_data' / 'chaines.csv'
CHAINES_DATA_AVAILABLE = False
CHAINES_REGEX = None
CHAINES_LOOKUP: dict[str, str] = {}

EUROPE_COUNTRY_CHOICES = [
    ('', 'Auto / non précisé'),
    ('FR', 'France'),
    ('IT', 'Italie'),
    ('ES', 'Espagne'),
    ('DE', 'Allemagne'),
    ('BE', 'Belgique'),
    ('NL', 'Pays-Bas'),
    ('GB', 'Royaume-Uni'),
    ('PT', 'Portugal'),
]

COUNTRY_NAME_TO_CODE = {
    'france': 'FR', 'fr': 'FR',
    'italy': 'IT', 'italia': 'IT', 'italie': 'IT', 'it': 'IT',
    'spain': 'ES', 'espana': 'ES', 'españa': 'ES', 'espagne': 'ES', 'es': 'ES',
    'germany': 'DE', 'deutschland': 'DE', 'allemagne': 'DE', 'de': 'DE',
    'belgium': 'BE', 'belgique': 'BE', 'belgie': 'BE', 'belgië': 'BE', 'be': 'BE',
    'netherlands': 'NL', 'pays bas': 'NL', 'pays-bas': 'NL', 'nederland': 'NL', 'holland': 'NL', 'nl': 'NL',
    'united kingdom': 'GB', 'uk': 'GB', 'gb': 'GB', 'great britain': 'GB', 'royaume uni': 'GB', 'royaume-uni': 'GB', 'england': 'GB',
    'portugal': 'PT', 'pt': 'PT',
}

COUNTRY_PROFILES = {
    'FR': {
        'street_types': ['rue', 'avenue', 'av', 'boulevard', 'bd', 'route', 'rte', 'chemin', 'allée', 'allee', 'place', 'impasse', 'quai', 'cours', 'route nationale', 'route departementale', 'parking', 'zone', 'zac', 'zi', 'zc'],
        'postcode_regex': r'^\d{5}$',
        'legal_id_type': 'siret_or_siren',
        'legal_prefixes': ['fr'],
        'number_suffixes': ['bis', 'ter', 'quater', 'quinquies', 'sexies'],
        'company_suffixes': ['sarl', 'sas', 'sa', 'eurl', 'sasu', 'scop'],
    },
    'IT': {
        'street_types': ['via', 'viale', 'piazza', 'corso', 'largo', 'strada', 'vicolo', 'piazzale', 'contrada', 'lungomare', 'vicolo', 'galleria'],
        'postcode_regex': r'^\d{5}$',
        'legal_id_type': 'partita_iva_or_codice_fiscale',
        'legal_prefixes': ['it'],
        'number_suffixes': ['bis', 'ter'],
        'company_suffixes': ['srl', 'spa', 'snc', 'sas', 'societa', 'società'],
    },
    'ES': {
        'street_types': ['calle', 'avenida', 'avda', 'plaza', 'paseo', 'camino', 'carretera', 'ronda', 'carrer', 'travessia', 'travesia'],
        'postcode_regex': r'^\d{5}$',
        'legal_id_type': 'nif_or_cif',
        'legal_prefixes': ['es'],
        'number_suffixes': ['bis'],
        'company_suffixes': ['sl', 'slu', 'sa', 'scoop'],
    },
    'DE': {
        'street_types': ['strasse', 'straße', 'platz', 'weg', 'allee', 'chaussee', 'ring', 'ufer', 'damm', 'gasse', 'markt'],
        'postcode_regex': r'^\d{5}$',
        'legal_id_type': 'vat_or_company_id',
        'legal_prefixes': ['de'],
        'number_suffixes': [],
        'company_suffixes': ['gmbh', 'ag', 'ug', 'kg', 'ohg'],
    },
    'BE': {
        'street_types': ['rue', 'avenue', 'chaussée', 'chaussee', 'boulevard', 'place', 'steenweg', 'straat', 'laan', 'plein'],
        'postcode_regex': r'^\d{4}$',
        'legal_id_type': 'vat_or_enterprise_number',
        'legal_prefixes': ['be'],
        'number_suffixes': ['bis'],
        'company_suffixes': ['sprl', 'srl', 'sa', 'bv', 'nv'],
    },
    'NL': {
        'street_types': ['straat', 'laan', 'plein', 'weg', 'markt', 'kade', 'hof', 'singel'],
        'postcode_regex': r'^\d{4}\s?[A-Z]{2}$',
        'legal_id_type': 'kvk_or_vat',
        'legal_prefixes': ['nl'],
        'number_suffixes': [],
        'company_suffixes': ['bv', 'nv', 'vof'],
    },
    'GB': {
        'street_types': ['street', 'st', 'road', 'rd', 'avenue', 'ave', 'lane', 'ln', 'close', 'court', 'way', 'drive', 'dr', 'high street'],
        'postcode_regex': r'^[A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2}$',
        'legal_id_type': 'company_or_vat',
        'legal_prefixes': ['gb', 'uk'],
        'number_suffixes': [],
        'company_suffixes': ['ltd', 'limited', 'plc', 'llp'],
    },
    'PT': {
        'street_types': ['rua', 'avenida', 'av', 'travessa', 'largo', 'praça', 'praca', 'estrada', 'alameda'],
        'postcode_regex': r'^\d{4}-?\d{3}$',
        'legal_id_type': 'nif',
        'legal_prefixes': ['pt'],
        'number_suffixes': [],
        'company_suffixes': ['lda', 'sa'],
    },
}
DEFAULT_COUNTRY_CODE = 'FR'
SUPPORTED_COUNTRY_CODES = [code for code, _ in EUROPE_COUNTRY_CHOICES if code]

CANONICAL_MAPPING_FIELDS = [
    'id', 'name', 'address', 'zipcode', 'city', 'country', 'legal_id',
    'lat', 'lng', 'hexa_gmap', 'phone_gmap', 'social_link_gmap',
]
REQUIRED_MATCHCODE_FIELDS = {'address', 'zipcode', 'city'}
COLUMN_ALIASES = {
    'id': ['id', 'identifier', 'identifiant', 'outlet_id', 'store_id', 'restaurant_id'],
    'name': ['name', 'nom', 'raison sociale', 'enseigne', 'outlet_name', 'store_name', 'ragione sociale'],
    'address': ['address', 'adresse', 'adresse1', 'street', 'rue', 'addr', 'full_address', 'indirizzo', 'direccion', 'dirección', 'strasse', 'straat'],
    'zipcode': ['zipcode', 'zip', 'postal_code', 'code_postal', 'cp', 'post_code', 'cap', 'postcode', 'plz'],
    'city': ['city', 'ville', 'commune', 'town', 'citta', 'città', 'ciudad', 'stadt', 'gemeente'],
    'country': ['country', 'pays', 'country_code', 'nation', 'paese', 'pais', 'país', 'land'],
    'legal_id': ['legal_id', 'siret', 'siren', 'vat', 'vat_number', 'partita_iva', 'piva', 'codice_fiscale', 'nif', 'cif', 'btw', 'kvk', 'company_number', 'ust_id', 'ustid'],
    'lat': ['lat', 'latitude'],
    'lng': ['lng', 'lon', 'long', 'longitude'],
    'hexa_gmap': ['hexa_gmap', 'hexa', 'hexa_code'],
    'phone_gmap': ['phone_gmap', 'phone', 'telephone', 'tel', 'mobile'],
    'social_link_gmap': ['social_link_gmap', 'website', 'web', 'site', 'url', 'social_link'],
}

REFERENCE_COLUMNS = {
    'id', 'hexa', 'name', 'address', 'zipcode', 'city', 'country', 'lat', 'lng',
    'vat', 'siren', 'siret', 'phone', 'email', 'website', 'voie', 'legal_id', 'legal_id_type',
    'num_voie', 'accessibility_gmap', 'activities_gmap', 'activity_gmap',
    'address_gmap', 'address_comp_gmap', 'advice_gmap', 'all_bookings_gmap',
    'all_deliveries_gmap', 'all_services_gmap', 'amenities_gmap',
    'atmosphere_gmap', 'business_status_gmap', 'cid_gmap', 'city_gmap',
    'code_plus_gmap', 'code_plus_city_gmap', 'country_gmap',
    'country_code_gmap', 'crowd_gmap', 'currency_gmap', 'delivery_gmap',
    'department_gmap', 'dining_options_gmap', 'full_address_gmap',
    'geocode_gmap', 'geoid_gmap', 'google_link_gmap', 'hexa_gmap',
    'hexa_link_gmap', 'info_link_gmap', 'label_gmap',
    'last_review_author_id_gmap', 'last_review_author_name_gmap',
    'last_review_date_gmap', 'lat_gmap', 'lng_gmap', 'name_gmap',
    'num_voie_gmap', 'offerings_gmap', 'outlet_description_gmap',
    'outlet_info_gmap', 'outlet_logo_gmap', 'owner_id_gmap',
    'owner_link_gmap', 'owner_name_gmap', 'payments_gmap', 'phone_gmap',
    'photo_gmap', 'place_id_gmap', 'planning_gmap', 'postal_code_gmap',
    'price_gmap', 'rate_gmap', 'region_gmap', 'social_link_gmap',
    'takeaway_gmap', 'voie_gmap', 'web_gmap', 'web_in_gmap',
    'week_schedule_gmap', 'zipcode_gmap', 'score_name_gmap',
    'score_address_gmap', 'score_city_gmap', 'score_zipcode_gmap',
    'score_num_voie', 'score_voie', 'distance_gmap', 'anomalie_distance',
    'gmap_automatch', 'geocode', 'matchcode', 'chaine'
}

COLUMNS_TO_KEEP = {
    'id', 'name', 'address', 'zipcode', 'city', 'country', 'legal_id', 'lat', 'lng',
    'hexa_gmap', 'phone_gmap', 'social_link_gmap'
}

PREFERRED_OUTPUT_ORDER = [
    'id', 'name', 'address', 'zipcode', 'city', 'country',
    'legal_id_type', 'legal_id',
    'chaine', 'matchcode', 'voie', 'num_voie',
    'lat', 'lng', 'hexa', 'phone', 'website',
]


def _build_stopword_pattern(street_types: list[str]) -> re.Pattern:
    sorted_words = sorted(set(street_types), key=len, reverse=True)
    escaped = '|'.join(re.escape(word) for word in sorted_words)
    return re.compile(rf'\b({escaped}|[a-zA-Z]\.)\b', flags=re.IGNORECASE)


COUNTRY_STOPWORD_PATTERNS = {
    code: _build_stopword_pattern(profile['street_types'])
    for code, profile in COUNTRY_PROFILES.items()
}


def _noop_progress(percent: int, message: str) -> None:
    return None


def _noop_log(message: str) -> None:
    return None


def normalized_label(value: str) -> str:
    return slugify(str(value or ''), separator='_')


def normalize_country_code(value: str | None) -> str:
    raw = slugify(str(value or ''), separator=' ')
    raw = re.sub(r'\s+', ' ', raw).strip()
    if not raw:
        return DEFAULT_COUNTRY_CODE
    if raw.upper() in COUNTRY_PROFILES:
        return raw.upper()
    return COUNTRY_NAME_TO_CODE.get(raw, DEFAULT_COUNTRY_CODE)


def country_profile(country_code: str | None) -> dict:
    return COUNTRY_PROFILES.get(normalize_country_code(country_code), COUNTRY_PROFILES[DEFAULT_COUNTRY_CODE])


def suggest_column_mapping(columns: list[str]) -> dict[str, str]:
    normalized_to_original = {normalized_label(col): col for col in columns}
    suggestions: dict[str, str] = {}
    used_sources: set[str] = set()
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            candidate = normalized_to_original.get(normalized_label(alias))
            if candidate and candidate not in used_sources:
                suggestions[target] = candidate
                used_sources.add(candidate)
                break
        if target in suggestions:
            continue
        for norm, original in normalized_to_original.items():
            if original in used_sources:
                continue
            if any(normalized_label(alias) in norm or norm in normalized_label(alias) for alias in aliases):
                suggestions[target] = original
                used_sources.add(original)
                break
    return suggestions


def detect_header_row(preview_rows: list[list[str]]) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(preview_rows):
        non_empty = [str(cell).strip() for cell in row if str(cell).strip()]
        unique_count = len(set(non_empty))
        score = (len(non_empty) * 10) + unique_count
        if score > best_score:
            best_score = score
            best_index = idx
    return best_index


def _sample_validation_warnings(ws, detected_columns: list[str], suggestions: dict[str, str], header_index: int) -> list[str]:
    warnings: list[str] = []
    column_positions = {col: idx for idx, col in enumerate(detected_columns)}
    sample_rows = []
    data_start = header_index + 2
    data_end = min(ws.max_row, data_start + 24)
    for row in ws.iter_rows(min_row=data_start, max_row=data_end, values_only=True):
        sample_rows.append(['' if value is None else str(value).strip() for value in row[: max(len(detected_columns), 20)]])

    def emptiness_ratio(source_col: str) -> float:
        idx = column_positions.get(source_col)
        if idx is None or not sample_rows:
            return 0.0
        empties = 0
        total = 0
        for row in sample_rows:
            total += 1
            value = row[idx] if idx < len(row) else ''
            if not str(value).strip():
                empties += 1
        return empties / total if total else 0.0

    for required_field in sorted(REQUIRED_MATCHCODE_FIELDS):
        source_col = suggestions.get(required_field)
        if not source_col:
            continue
        ratio = emptiness_ratio(source_col)
        if ratio >= 0.5:
            warnings.append(
                f"La colonne suggérée '{source_col}' pour {required_field} semble vide à {int(ratio * 100)}% sur l’échantillon."
            )

    lat_col = suggestions.get('lat')
    lng_col = suggestions.get('lng')
    if lat_col and lng_col and sample_rows:
        lat_idx = column_positions.get(lat_col)
        lng_idx = column_positions.get(lng_col)
        invalid_points = 0
        checked_points = 0
        for row in sample_rows:
            try:
                lat_raw = row[lat_idx] if lat_idx is not None and lat_idx < len(row) else ''
                lng_raw = row[lng_idx] if lng_idx is not None and lng_idx < len(row) else ''
                if not lat_raw and not lng_raw:
                    continue
                checked_points += 1
                lat_val = float(str(lat_raw).replace(',', '.'))
                lng_val = float(str(lng_raw).replace(',', '.'))
                if not (-90 <= lat_val <= 90 and -180 <= lng_val <= 180):
                    invalid_points += 1
            except Exception:
                invalid_points += 1
                checked_points += 1
        if checked_points and invalid_points:
            warnings.append(
                f"{invalid_points} coordonnées lat/lng semblent invalides sur {checked_points} lignes de l’échantillon."
            )

    legal_col = suggestions.get('legal_id')
    if legal_col and sample_rows:
        ratio = emptiness_ratio(legal_col)
        if ratio >= 0.6:
            warnings.append(
                f"La colonne suggérée '{legal_col}' pour legal_id semble vide à {int(ratio * 100)}% sur l’échantillon."
            )

    return warnings


def inspect_excel_workbook(uploaded_file) -> dict:
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheets = []
    for ws in workbook.worksheets[:10]:
        rows = []
        preview_limit = min(ws.max_row, 20)
        max_preview_columns = min(ws.max_column, 30)
        for row in ws.iter_rows(min_row=1, max_row=preview_limit, values_only=True):
            rows.append(['' if value is None else str(value)[:120] for value in (row or [])[:max_preview_columns]])
        header_index = detect_header_row(rows[:8]) if rows else 0
        detected_columns = rows[header_index] if rows else []
        detected_columns = [str(col).strip() for col in detected_columns if str(col).strip()]
        suggestions = suggest_column_mapping(detected_columns)
        sheets.append({
            'name': ws.title,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'preview': rows[:20],
            'detected_header_row': header_index + 1,
            'detected_columns': detected_columns,
            'mapping_suggestions': suggestions,
            'missing_required_for_matchcode': sorted(REQUIRED_MATCHCODE_FIELDS - set(suggestions.keys())),
            'validation_warnings': _sample_validation_warnings(ws, detected_columns, suggestions, header_index),
        })
    workbook.close()
    return sheets


def load_chaines_data() -> tuple[bool, str | None]:
    global CHAINES_DATA_AVAILABLE, CHAINES_REGEX, CHAINES_LOOKUP
    if CHAINES_DATA_AVAILABLE:
        return True, None
    if not CHAINES_CSV_PATH.is_file():
        return False, f"Fichier chaînes absent : {CHAINES_CSV_PATH.name}. La colonne 'chaine' restera vide."

    try:
        df = pd.read_csv(CHAINES_CSV_PATH)
        if 'name' not in df.columns or 'keyword' not in df.columns:
            return False, "Le fichier chaines.csv doit contenir les colonnes 'name' et 'keyword'."

        df = df.dropna(subset=['keyword', 'name']).copy()
        df['slug_keyword'] = df['keyword'].apply(lambda x: slugify(str(x)))
        df = df.dropna(subset=['slug_keyword'])
        df = df.drop_duplicates(subset=['slug_keyword'])
        df['keyword_len'] = df['slug_keyword'].str.len()
        df = df.sort_values(by='keyword_len', ascending=False)

        CHAINES_LOOKUP = pd.Series(df.name.values, index=df.slug_keyword).to_dict()
        bounded_keywords = [r'\b' + re.escape(k) + r'\b' for k in df['slug_keyword']]
        CHAINES_REGEX = re.compile('|'.join(bounded_keywords)) if bounded_keywords else None
        CHAINES_DATA_AVAILABLE = True
        return True, f"Chaînes chargées : {len(df)} mots-clés"
    except Exception as exc:
        return False, f"Impossible de charger chaines.csv : {exc}"


def contains_at_least_one_number(value) -> bool:
    if pd.isna(value):
        return False
    return bool(re.search(r'\d', str(value)))


def normalize_postcode(value, country_code: str | None) -> str:
    if pd.isna(value) or value is None:
        return ''
    country_code = normalize_country_code(country_code)
    raw = str(value).strip().upper()
    raw = re.sub(r'\.0$', '', raw)
    raw = re.sub(r'[^A-Z0-9\- ]', '', raw)
    compact = re.sub(r'\s+', '', raw)
    if country_code in {'FR', 'IT', 'ES', 'DE'}:
        digits = re.sub(r'\D', '', compact)
        return digits.zfill(5) if digits and len(digits) <= 5 else digits[:5]
    if country_code == 'BE':
        digits = re.sub(r'\D', '', compact)
        return digits.zfill(4) if digits and len(digits) <= 4 else digits[:4]
    if country_code == 'NL':
        m = re.match(r'^(\d{4})([A-Z]{2})$', compact)
        return f"{m.group(1)} {m.group(2)}" if m else compact
    if country_code == 'GB':
        compact = compact.replace(' ', '')
        return f"{compact[:-3]} {compact[-3:]}".strip() if len(compact) > 3 else compact
    if country_code == 'PT':
        digits = re.sub(r'\D', '', compact)
        return f"{digits[:4]}-{digits[4:7]}" if len(digits) >= 7 else digits
    return compact


def infer_legal_id_type(value, country_code: str | None) -> str:
    country_code = normalize_country_code(country_code)
    cleaned = normalize_legal_id(value, country_code)
    if not cleaned:
        return ''
    if country_code == 'FR':
        if cleaned.isdigit() and len(cleaned) == 14:
            return 'siret'
        if cleaned.isdigit() and len(cleaned) == 9:
            return 'siren'
    if country_code == 'IT':
        if cleaned.isdigit() and len(cleaned) == 11:
            return 'partita_iva'
        if len(cleaned) == 16:
            return 'codice_fiscale'
    if country_code == 'ES' and len(cleaned) == 9:
        return 'nif_cif'
    if country_code == 'BE' and cleaned.isdigit() and len(cleaned) == 10:
        return 'enterprise_number'
    if country_code == 'NL':
        if cleaned.isdigit() and len(cleaned) == 8:
            return 'kvk'
        return 'btw_or_kvk'
    if country_code == 'GB':
        if cleaned.isdigit() and len(cleaned) == 8:
            return 'company_number'
        return 'vat_or_company_number'
    if country_code == 'PT' and cleaned.isdigit() and len(cleaned) == 9:
        return 'nif'
    return country_profile(country_code).get('legal_id_type', 'legal_id')


def normalize_legal_id(value, country_code: str | None) -> str:
    if pd.isna(value) or value is None:
        return ''
    country_code = normalize_country_code(country_code)
    raw = str(value).upper().strip()
    raw = re.sub(r'[^A-Z0-9]', '', raw)
    for prefix in country_profile(country_code).get('legal_prefixes', []):
        prefix_upper = prefix.upper()
        if raw.startswith(prefix_upper) and len(raw) > len(prefix_upper):
            raw = raw[len(prefix_upper):]
            break
    if country_code == 'FR':
        digits = re.sub(r'\D', '', raw)
        if len(digits) >= 14:
            return digits[:14]
        if len(digits) >= 9:
            return digits[:9]
        return digits
    if country_code == 'IT':
        if raw.isdigit() and len(raw) >= 11:
            return raw[:11]
        return raw[:16]
    if country_code == 'ES':
        return raw[:9]
    if country_code == 'BE':
        digits = re.sub(r'\D', '', raw)
        return digits[:10]
    if country_code == 'NL':
        if raw.isdigit():
            return raw[:8]
        return raw[:14]
    if country_code == 'GB':
        return raw[:12]
    if country_code == 'PT':
        digits = re.sub(r'\D', '', raw)
        return digits[:9]
    return raw


def _find_raw_num_voie_match(address, country_code: str | None = None) -> str | None:
    if pd.isna(address):
        return None
    address = str(address)
    patterns = [
        r'\b(\d+[A-Z]?)\b',
        r'\b(\d+[/\-]\d+)\b',
        r'\b([A-Z]?\d+[A-Z]?)\b',
    ]
    for pattern in patterns:
        match = re.search(pattern, address, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def detect_num_voie(address, country_code: str | None = None) -> str | None:
    raw_match = _find_raw_num_voie_match(address, country_code)
    if raw_match:
        return re.sub(r'[\s_]', '', raw_match).upper()
    return None


def detect_voie(address, country_code: str | None = None) -> str | None:
    if pd.isna(address):
        return None
    country_code = normalize_country_code(country_code)
    address_str = str(address)
    raw_num_voie = _find_raw_num_voie_match(address_str, country_code)
    if raw_num_voie:
        initial_voie = address_str.replace(raw_num_voie, '', 1).strip(' ,-/')
    else:
        initial_voie = address_str.strip()
    if not initial_voie:
        return initial_voie
    cleaned_voie = COUNTRY_STOPWORD_PATTERNS.get(country_code, COUNTRY_STOPWORD_PATTERNS[DEFAULT_COUNTRY_CODE]).sub('', initial_voie)
    processed_voie = slugify(cleaned_voie, separator=' ').upper()
    processed_voie = ' '.join(processed_voie.split())
    if not processed_voie:
        return slugify(initial_voie, separator=' ').upper()
    return processed_voie


def make_matchcode(address, zipcode, country_code: str | None = None) -> str | None:
    if pd.isna(address) or pd.isna(zipcode) or not address or not zipcode:
        return None
    country_code = normalize_country_code(country_code)
    zipcode = normalize_postcode(zipcode, country_code)
    address_proc = str(address).replace('-', ' ').replace("'", ' ')
    words = [word for word in address_proc.split() if word]
    if not words:
        return None

    number_candidate, last_word = None, None
    first_word, last_word = words[0], words[-1]
    if contains_at_least_one_number(first_word):
        number_candidate = first_word
    elif contains_at_least_one_number(last_word):
        number_candidate = last_word

    if number_candidate and number_candidate == words[-1]:
        temp_address_words = words[:-1]
        last_word = temp_address_words[-1] if temp_address_words else None

    if not number_candidate and len(words) > 1 and contains_at_least_one_number(words[-2]):
        number_candidate = words[-2]

    if not number_candidate:
        match = re.search(r'\b(\d+[A-Z]?)\b', address_proc, flags=re.IGNORECASE)
        if match:
            number_candidate = match.group(1)
            parts = address_proc.split(number_candidate, 1)
            if parts[0].strip():
                last_word = parts[0].strip().split()[-1]

    if not number_candidate or not last_word:
        return None

    final_number = number_candidate
    addr_slug = slugify(address_proc, separator=' ')
    for suffix in country_profile(country_code).get('number_suffixes', []):
        if f"{slugify(final_number)} {suffix}" in addr_slug:
            final_number = f"{final_number}{suffix}"
            break

    return f"{country_code}-{zipcode}-{slugify(last_word)}-{slugify(final_number)}"


def find_chaine_local(name: str) -> str | None:
    if not CHAINES_DATA_AVAILABLE or pd.isna(name) or not CHAINES_REGEX:
        return None
    slug_name = slugify(str(name))
    if not slug_name:
        return None
    match = CHAINES_REGEX.search(slug_name)
    if match:
        return CHAINES_LOOKUP.get(match.group(0))
    return None


@dataclass(slots=True)
class NormalizerOptions:
    do_clean: bool = True
    do_matchcode: bool = True
    sheet_name: str | None = None
    column_mapping: dict[str, str] = field(default_factory=dict)
    country_code: str | None = None


class NormalizerService:
    def __init__(self, progress_callback: ProgressCallback | None = None, log_callback: LogCallback | None = None):
        self.progress_callback = progress_callback or _noop_progress
        self.log_callback = log_callback or _noop_log

    def _progress(self, percent: int, message: str) -> None:
        self.progress_callback(percent, message)

    def _log(self, message: str) -> None:
        self.log_callback(message)

    def run(self, input_path: str | Path, output_path: str | Path, options: NormalizerOptions) -> Path:
        input_path = Path(input_path)
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if not input_path.exists():
            raise FileNotFoundError(f"Le fichier n'existe pas : {input_path}")
        if input_path.suffix.lower() not in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
            raise ValueError('Le normalizer V1 web supporte uniquement les fichiers Excel .xlsx/.xlsm/.xltx/.xltm.')

        chosen_country = normalize_country_code(options.country_code)
        profile = country_profile(chosen_country)
        self._log(f"🌍 Profil pays utilisé : {chosen_country} ({dict(EUROPE_COUNTRY_CHOICES).get(chosen_country, chosen_country)})")
        self._log(f"🧾 Spécificités actives : adresses={len(profile['street_types'])} types de voie, legal_id={profile['legal_id_type']}")

        self._progress(5, 'Analyse du fichier Excel')
        xls = pd.ExcelFile(input_path)
        sheet_names = xls.sheet_names
        chosen_sheet = options.sheet_name.strip() if options.sheet_name else None
        if chosen_sheet and chosen_sheet not in sheet_names:
            raise ValueError(f"Onglet introuvable : {chosen_sheet}. Onglets disponibles : {', '.join(sheet_names)}")
        if not chosen_sheet:
            chosen_sheet = sheet_names[0]
            if len(sheet_names) > 1:
                self._log(f"ℹ️ Plusieurs onglets détectés : {', '.join(sheet_names)}")
                self._log(f"ℹ️ Aucun onglet fourni, utilisation du premier : {chosen_sheet}")

        self._log(f"📄 Onglet utilisé : {chosen_sheet}")
        df = pd.read_excel(xls, sheet_name=chosen_sheet)
        self._log(f"✓ Fichier chargé : {len(df)} lignes, {len(df.columns)} colonnes")
        self._log('🧾 Colonnes détectées : ' + ', '.join(map(str, df.columns.tolist()[:20])) + (' …' if len(df.columns) > 20 else ''))
        self._progress(15, 'Lecture du classeur terminée')

        if options.column_mapping:
            df = self._apply_column_mapping(df, options.column_mapping)

        if 'country' not in df.columns or df['country'].isna().all():
            df['country'] = chosen_country
            self._log(f"🌍 Colonne country injectée avec la valeur par défaut : {chosen_country}")
        else:
            df['country'] = df['country'].apply(normalize_country_code)
            self._log('🌍 Colonne country normalisée à partir du fichier source')

        if options.do_matchcode:
            missing = sorted(REQUIRED_MATCHCODE_FIELDS - set(df.columns))
            if missing:
                raise ValueError(
                    'Colonnes requises manquantes pour le matchcode : ' + ', '.join(missing) + '. '
                    'Renseigne le mapping de colonnes avant de relancer le job.'
                )

        if options.do_clean:
            df = self._perform_cleaning(df, chosen_country)

        if options.do_matchcode:
            df = self._perform_matchcode(df, chosen_country)

        df = self._reorder_output_columns(df)
        self._log('🧱 Ordre final des colonnes aligné avec le normalizer desktop')

        self._progress(95, 'Écriture du fichier résultat CSV')
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        self._log(f"✓ Fichier CSV sauvegardé : {output_path.name}")
        self._log(f"📊 Résumé : {len(df)} lignes, {len(df.columns)} colonnes")
        self._log('✅ V14 multi-country Europe : adresses, codes postaux et legal_id normalisés')
        self._progress(100, 'Traitement terminé')
        return output_path

    def _apply_column_mapping(self, df: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
        cleaned_mapping = {}
        for target, source in mapping.items():
            source = str(source or '').strip()
            if not source or source == '__ignore__':
                continue
            if source not in df.columns:
                raise ValueError(f"La colonne source '{source}' est introuvable pour le mapping '{target}'.")
            cleaned_mapping[target] = source

        duplicates = [src for src in cleaned_mapping.values() if list(cleaned_mapping.values()).count(src) > 1]
        if duplicates:
            raise ValueError('Une même colonne source ne peut pas être mappée plusieurs fois : ' + ', '.join(sorted(set(duplicates))))

        reverse_mapping = {source: target for target, source in cleaned_mapping.items() if source != target}
        if reverse_mapping:
            self._log('🧭 Mapping appliqué : ' + ', '.join(f"{src} → {dst}" for src, dst in reverse_mapping.items()))
            df = df.rename(columns=reverse_mapping)
        return df

    def _perform_cleaning(self, df: pd.DataFrame, chosen_country: str) -> pd.DataFrame:
        self._log('--- Nettoyage des données ---')
        self._progress(25, 'Nettoyage des colonnes')

        columns_to_remove = []
        custom_columns = []
        for col in df.columns:
            if col not in COLUMNS_TO_KEEP:
                if col in REFERENCE_COLUMNS or 'gmap' in col.lower():
                    columns_to_remove.append(col)
                else:
                    custom_columns.append(col)

        df = df.drop(columns=columns_to_remove, errors='ignore')
        if columns_to_remove:
            self._log(f"✓ {len(columns_to_remove)} colonnes supprimées")
            self._log('🗑️ Colonnes supprimées : ' + ', '.join(columns_to_remove[:20]) + (' …' if len(columns_to_remove) > 20 else ''))
        if custom_columns:
            self._log(f"✓ {len(custom_columns)} colonnes personnalisées conservées : {', '.join(custom_columns)}")

        old_columns_to_drop = []
        if 'hexa' in df.columns and 'hexa_gmap' in df.columns:
            old_columns_to_drop.append('hexa')
        if 'phone' in df.columns and 'phone_gmap' in df.columns:
            old_columns_to_drop.append('phone')
        if 'website' in df.columns and 'social_link_gmap' in df.columns:
            old_columns_to_drop.append('website')
        if old_columns_to_drop:
            df = df.drop(columns=old_columns_to_drop, errors='ignore')
            self._log(f"✓ Anciennes colonnes supprimées avant renommage : {', '.join(old_columns_to_drop)}")

        rename_mapping = {
            'hexa_gmap': 'hexa',
            'phone_gmap': 'phone',
            'social_link_gmap': 'website',
        }
        df = df.rename(columns=rename_mapping)

        for col in ['name', 'address', 'city']:
            if col in df.columns:
                df[col] = df[col].fillna('').astype(str).str.strip()
        if 'country' in df.columns:
            df['country'] = df['country'].fillna(chosen_country).apply(normalize_country_code)
        if 'zipcode' in df.columns:
            df['zipcode'] = df.apply(lambda row: normalize_postcode(row.get('zipcode'), row.get('country') or chosen_country), axis=1)
        if 'legal_id' in df.columns:
            df['legal_id'] = df.apply(lambda row: normalize_legal_id(row.get('legal_id'), row.get('country') or chosen_country), axis=1)
            df['legal_id_type'] = df.apply(lambda row: infer_legal_id_type(row.get('legal_id'), row.get('country') or chosen_country), axis=1)
            legal_hits = int(df['legal_id'].astype(str).str.len().gt(0).sum())
            self._log(f"🆔 legal_id normalisé ({legal_hits}/{len(df)})")
        else:
            df['legal_id_type'] = ''

        self._log('✓ Colonnes renommées et harmonisées avec succès')
        self._log('🧽 Colonnes après nettoyage : ' + ', '.join(map(str, df.columns.tolist()[:20])) + (' …' if len(df.columns) > 20 else ''))
        self._progress(40, 'Nettoyage terminé')
        return df

    def _perform_matchcode(self, df: pd.DataFrame, chosen_country: str) -> pd.DataFrame:
        self._log('--- Génération des matchcodes ---')
        self._progress(50, 'Préparation des colonnes address/zipcode/city')

        for col in ['name', 'address', 'zipcode', 'city', 'country']:
            if col in df.columns:
                df[col] = df[col].fillna('').astype(str)
        if 'country' not in df.columns:
            df['country'] = chosen_country
        df['country'] = df['country'].apply(normalize_country_code)
        if 'zipcode' in df.columns:
            df['zipcode'] = df.apply(lambda row: normalize_postcode(row['zipcode'], row.get('country') or chosen_country), axis=1)

        self._progress(60, 'Calcul des colonnes num_voie et voie')
        df['num_voie'] = df.apply(lambda row: detect_num_voie(row['address'], row.get('country') or chosen_country), axis=1)
        df['voie'] = df.apply(lambda row: detect_voie(row['address'], row.get('country') or chosen_country), axis=1)

        self._progress(72, 'Calcul des matchcodes')
        df['matchcode'] = df.apply(lambda row: make_matchcode(row['address'], row['zipcode'], row.get('country') or chosen_country), axis=1)
        matchcode_count = int(df['matchcode'].notna().sum())
        self._log(f'✓ Matchcodes générés ({matchcode_count}/{len(df)})')

        chains_loaded, chains_message = load_chaines_data()
        if chains_message:
            self._log(('✓ ' if chains_loaded else 'ℹ️ ') + chains_message)
        self._progress(82, 'Recherche des chaînes')
        if 'name' in df.columns and chains_loaded:
            df['chaine'] = df['name'].apply(find_chaine_local)
            chaine_hits = int(df['chaine'].notna().sum())
            self._log(f'✓ Recherche des chaînes terminée ({chaine_hits} correspondances)')
        else:
            df['chaine'] = None
            self._log("ℹ️ Colonne 'chaine' laissée vide")

        cols = df.columns.tolist()
        if 'city' in cols:
            city_index = cols.index('city')
            new_cols_order = ['country', 'legal_id_type', 'legal_id', 'chaine', 'matchcode', 'voie', 'num_voie']
            for column in new_cols_order:
                if column in cols:
                    cols.remove(column)
            for column in reversed(new_cols_order):
                if column in df.columns:
                    cols.insert(city_index + 1, column)
            df = df[cols]

        self._progress(90, 'Réorganisation des colonnes terminée')
        return df

    def _reorder_output_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        preferred = [col for col in PREFERRED_OUTPUT_ORDER if col in df.columns]
        remaining = [col for col in df.columns if col not in preferred]
        return df[preferred + remaining]
