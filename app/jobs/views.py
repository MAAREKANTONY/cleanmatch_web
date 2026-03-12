from pathlib import Path

from django.contrib import messages
from django.db.models import Count
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import load_workbook
from rest_framework.decorators import api_view
from rest_framework.response import Response

from .forms import JobCreateForm
from .models import Job
from .serializers import JobSerializer
from .services import JobService
from .tasks import run_uploaded_job


def home(request):
    jobs = Job.objects.all()

    job_type = request.GET.get('job_type', '').strip()
    status = request.GET.get('status', '').strip()
    q = request.GET.get('q', '').strip()

    if job_type:
        jobs = jobs.filter(job_type=job_type)
    if status:
        jobs = jobs.filter(status=status)
    if q:
        jobs = jobs.filter(input_file_1__icontains=q)

    stats_raw = Job.objects.values('status').annotate(total=Count('id'))
    stats = {row['status']: row['total'] for row in stats_raw}
    context = {
        'jobs': jobs[:30],
        'stats': {
            'total': Job.objects.count(),
            'running': stats.get(Job.Status.RUNNING, 0),
            'queued': stats.get(Job.Status.QUEUED, 0),
            'success': stats.get(Job.Status.SUCCESS, 0),
            'failed': stats.get(Job.Status.FAILED, 0),
        },
        'filters': {
            'job_type': job_type,
            'status': status,
            'q': q,
        },
        'job_type_choices': Job.JobType.choices,
        'status_choices': Job.Status.choices,
    }
    return render(request, 'jobs/home.html', context)


def create_job(request):
    if request.method == 'POST':
        form = JobCreateForm(request.POST, request.FILES)
        if form.is_valid():
            parameters = {
                'mode': 'uploaded',
                'filename_1': form.cleaned_data['input_file_1'].name,
                'filename_2': form.cleaned_data['input_file_2'].name if form.cleaned_data.get('input_file_2') else None,
            }
            if form.cleaned_data['job_type'] == Job.JobType.NORMALIZER:
                parameters.update({
                    'do_clean': form.cleaned_data['normalizer_do_clean'],
                    'do_matchcode': form.cleaned_data['normalizer_do_matchcode'],
                    'sheet_name': form.cleaned_data['normalizer_sheet_name'].strip(),
                })

            job = Job.objects.create(
                job_type=form.cleaned_data['job_type'],
                status=Job.Status.PENDING,
                progress_message='Job créé',
                parameters_json=parameters,
                input_file_1=form.cleaned_data['input_file_1'],
                input_file_2=form.cleaned_data.get('input_file_2') or None,
            )
            async_result = run_uploaded_job.delay(str(job.id))
            JobService.mark_queued(job, async_result.id)
            messages.success(request, f'Job {job.id} créé et envoyé au worker.')
            return redirect('jobs:detail', job_id=job.id)
    else:
        form = JobCreateForm()

    return render(request, 'jobs/new.html', {'form': form})


def job_detail(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    return render(request, 'jobs/job_detail.html', {'job': job})


@api_view(['GET'])
def api_job_detail(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    serializer = JobSerializer(job, context={'request': request})
    return Response(serializer.data)


def inspect_excel(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée.'}, status=405)

    uploaded = request.FILES.get('input_file_1')
    if not uploaded:
        return JsonResponse({'error': 'Aucun fichier fourni.'}, status=400)

    filename = uploaded.name.lower()
    allowed_ext = {'.xlsx', '.xlsm', '.xltx', '.xltm'}
    if not any(filename.endswith(ext) for ext in allowed_ext):
        return JsonResponse({'error': 'Inspection disponible uniquement pour les fichiers Excel.'}, status=400)

    try:
        workbook = load_workbook(uploaded, read_only=True, data_only=True)
        sheets = []
        for ws in workbook.worksheets[:10]:
            rows = ws.iter_rows(min_row=1, max_row=3, values_only=True)
            preview = []
            for row in rows:
                preview.append([
                    '' if value is None else str(value)[:80]
                    for value in (row or [])[:12]
                ])
            sheets.append({
                'name': ws.title,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'preview': preview,
            })
        workbook.close()
    except Exception as exc:
        return JsonResponse({'error': f'Impossible de lire le fichier Excel : {exc}'}, status=400)

    return JsonResponse({
        'filename': Path(uploaded.name).name,
        'sheets': sheets,
    })
